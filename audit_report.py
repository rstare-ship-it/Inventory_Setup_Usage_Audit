#!/usr/bin/env python3
"""
Inventory Setup / Usage Audit

For tenants already using the inventory module. Parses Snowflake query results
(POs, invoice materials, replenishment, assessment) and produces an audit report.
Checks adapted from Inventory Readiness Check; thresholds are tweakable.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path


def _sanitize_folder_name(s: str, max_len: int = 80) -> str:
    """Safe folder name: alphanumeric, underscore, hyphen only; collapse runs; strip."""
    if not s or not isinstance(s, str):
        return "report"
    name = re.sub(r"[^\w\-]+", "_", s.strip())
    name = re.sub(r"_+", "_", name).strip("_")
    return (name[:max_len] if len(name) > max_len else name) or "report"

# Load .env from project directory so SNOWFLAKE_* are set when using --from-snowflake
try:
    from dotenv import load_dotenv
    _script_dir = Path(__file__).resolve().parent
    load_dotenv(_script_dir / ".env")
except ImportError:
    pass

# Result set keys (combined query returns 10 rows; first row = tenant_info for header; pricebook for readiness).
RESULT_KEYS = [
    "tenant_info",
    "pricebook",
    "purchase_orders_summary",
    "invoice_materials",
    "replenishment_summary",
    "returns_summary",
    "assessment_data",
    "inventory_settings",
    "setup_data",
    "usage_checks",
]

# PO status (app model): 0=Pending, 1=Exported, 2=Sent, 3=PartiallyReceived, 4=Received, 5=Canceled, 6=PendingApproval
PENDING_STATUS_WAITING_TO_SEND = 0
PENDING_STATUSES = (0, 6, 2, 3)
COMPLETED_STATUSES = (4, 1)

# --- Tweakable thresholds (audit) ---
# POs: total pending is relative to trucks (allowed = half of trucks); absolute fallback when no trucks
TOTAL_PENDING_ABSOLUTE_RED = 50
PENDING_POS_ALLOWED_PER_TRUCK = 0.5   # allowed pending = this fraction of truck count (e.g. half)
# Pending as % of allowed (4 levels, lower better): ≤25% Good, 25-40% Okay, 40-60% Needs review, >60% Needs attention
PENDING_PCT_OF_ALLOWED_GOOD_MAX = 25
PENDING_PCT_OF_ALLOWED_OKAY_MAX = 40
PENDING_PCT_OF_ALLOWED_REVIEW_MAX = 60
# Pending over 90 days as % of allowed (4 levels, lower better): <5% Good, 5-10% Okay, 10-30% Needs review, >30% Needs attention
PENDING_OVER_90_PCT_GOOD_MAX = 5
PENDING_OVER_90_PCT_OKAY_MAX = 10
PENDING_OVER_90_PCT_REVIEW_MAX = 30
# PO line item: % of POs with multiple lines (4 levels, higher better): <40% red, 40-50% review, 50-60% yellow, >60% green
PO_MULTI_LINE_RED_MAX_PCT = 40
PO_MULTI_LINE_REVIEW_MAX_PCT = 50
PO_MULTI_LINE_YELLOW_MIN_PCT = 50
PO_MULTI_LINE_YELLOW_MAX_PCT = 60
# PO receive rate % (4 levels, higher better): ≥80% Good, 65-80% Okay, 50-65% Needs review, <50% Needs attention
PO_RECEIVE_RATE_GOOD_MIN_PCT = 80
PO_RECEIVE_RATE_OKAY_MIN_PCT = 65
PO_RECEIVE_RATE_REVIEW_MIN_PCT = 50
PO_PLACEHOLDER_MAJORITY = 0.5   # red if > 50% of PO line items have placeholder/generic/material
# Invoice: % of invoices with $>0 that have ≥1 IsInventory material line (4 levels): 85%+ Good, 75-85% Okay, 50-75% Needs review, <50% Needs attention
INVOICE_GT_ZERO_WITH_MATERIAL_GOOD_MIN_PCT = 85
INVOICE_GT_ZERO_WITH_MATERIAL_OKAY_MIN_PCT = 75
INVOICE_GT_ZERO_WITH_MATERIAL_REVIEW_MIN_PCT = 50
# Legacy placeholder/tech thresholds (still used if we add those checks back)
INVOICE_PLACEHOLDER_GREEN_MAX_PCT = 10
INVOICE_PLACEHOLDER_RED_MIN_PCT = 70
# Replenishment rate (4 levels, higher better): 80%+ Good, 70-80% Okay, 50-70% Needs review, <50% Needs attention
REPLENISHMENT_RATE_GOOD_MIN_PCT = 80
REPLENISHMENT_RATE_OKAY_MIN_PCT = 70
REPLENISHMENT_RATE_REVIEW_MIN_PCT = 50
# Invoice: technician-added % (4 levels, higher better): ≥60% Good, 50-60% Okay, 30-50% Needs review, <30% Needs attention
INVOICE_TECH_ADDED_GOOD_MIN_PCT = 60
INVOICE_TECH_ADDED_OKAY_MIN_PCT = 50
INVOICE_TECH_ADDED_REVIEW_MIN_PCT = 30
# Returns: status 2 = CreditReceived; pending % of returns in period (4 levels, lower better): ≤5% Good, 5-10% Okay, 10-20% Needs review, >20% Needs attention
RETURN_CREDIT_RECEIVED_STATUS = 2
RETURN_PENDING_PCT_GOOD_MAX = 5
RETURN_PENDING_PCT_OKAY_MAX = 10
RETURN_PENDING_PCT_REVIEW_MAX = 20
# Setup: trucks/warehouses/templates
TRUCK_TEMPLATE_PCT_MIN = 80   # at least 80% of (active) trucks must have an inventory template
WAREHOUSE_WITH_TEMPLATE_MIN = 1  # at least 1 (active) warehouse with template
TEMPLATE_MIN_ACTIVE_ITEMS = 20   # each template should have at least 20 active items
# Usage checks (red when exceeded)
USAGE_REPLENISHMENT_OLD_30D_MAX = 20   # red if > 20 replenishment requests older than 30 days
USAGE_TRANSFERS_OLD_14D_MAX = 20       # red if > 20 pending transfers older than 14 days
USAGE_REQUISITIONS_OLD_90D_MAX = 20     # red if >= 20 requisitions older than 90 days
USAGE_PAST_DUE_COUNTS_MAX = 5          # red if > 5 past due inventory counts
USAGE_DIRECT_ADJUSTMENTS_MAX = 20      # red if > 20 direct (type 0) adjustments in last 90 days
# Extended checks — invoice material lines IsInventory % (4 levels): 85%+ Good, 75-85% Okay, 50-75% Needs review, <50% Needs attention
INVOICE_MATERIAL_LINES_ISINVENTORY_GOOD_MIN_PCT = 85
INVOICE_MATERIAL_LINES_ISINVENTORY_OKAY_MIN_PCT = 75
INVOICE_MATERIAL_LINES_ISINVENTORY_REVIEW_MIN_PCT = 50
# Zero-cost IsInventory lines (4 levels, lower better): <3% Good, 3-5% Okay, 5-15% Needs review, >=15% Needs attention
INVOICE_ZERO_COST_GOOD_MAX_PCT = 3
INVOICE_ZERO_COST_OKAY_MAX_PCT = 5
INVOICE_ZERO_COST_REVIEW_MAX_PCT = 15
# Transfers: expect half of trucks to get at least one transfer per week over lookback
TRANSFER_EXPECTED_TRUCK_FRACTION = 0.5  # half of trucks
MATERIALS_NO_UOM_YELLOW_MAX_PCT = 10   # (informational only; not scored)
MATERIALS_NO_UOM_RED_MIN_PCT = 25      # (informational only; not scored)
MATERIALS_UNUSED_90D_YELLOW_MAX_PCT = 50   # yellow when > 50% of IsInventory materials had no invoice usage in 90 days
MATERIALS_UNUSED_90D_RED_MIN_PCT = 75      # red when > 75%
WAREHOUSES_NO_COUNT_90D_MAX = 0       # red when any warehouse has no completed count in last 90 days (0 = all must have one)

# Audit areas for scorecard (display order) — full inventory audit
AUDIT_AREAS = [
    "Pricebook & setup",
    "Purchasing",
    "Invoicing",
    "Technician usage",
    "Replenishment",
    "Returns",
    "Transfers",
    "Counts and adjustments",
    "Other",
]

# Areas for purchasing-only audit (when not yet live with inventory; no template/setup — inventory-only)
AUDIT_AREAS_PURCHASING = [
    "Pricebook",
    "Purchasing",
    "Invoicing",
    "Replenishment",
    "Returns",
]

# Areas that determine "ready for inventory implementation" (only these are listed when not ready)
READINESS_AREAS = ["Pricebook", "Purchasing", "Invoicing"]

# Pricebook readiness (from Inventory_Prepardness_Check): red if > this share of materials have $0 cost or Default Replenishment as primary
PRICEBOOK_RED_PCT = 0.25


def _readiness_areas_needing_work(findings_by_area: dict[str, dict[str, list[str]]]) -> list[str]:
    """Return list of READINESS_AREAS that have any red findings (for 'what to work on' message)."""
    return [a for a in READINESS_AREAS if findings_by_area.get(a, {}).get("red")]


def _format_areas_for_message(areas: list[str]) -> str:
    """Format area names for verdict message: 'A', 'A and B', or 'A, B, and C'."""
    if not areas:
        return ""
    if len(areas) == 1:
        return areas[0]
    if len(areas) == 2:
        return f"{areas[0]} and {areas[1]}"
    return f"{areas[0]}, {areas[1]}, and {areas[2]}"


# Go Live readiness: min share of materials marked IsInventory (same as full-audit template thresholds)
GO_LIVE_ITEMS_INVENTORY_MIN_PCT = 50  # at least half of materials marked for inventory


def _is_preparing_for_go_live(results: dict, is_live: bool) -> bool:
    """True when not live but inventory module on, has items marked IsInventory, and has templates assigned."""
    if is_live:
        return False
    cfg = results.get("readiness_config") or {}
    if not cfg.get("inventory_module_on"):
        return False
    isinv = results.get("isinventory_counts") or {}
    mat_isinv = _int(isinv.get("materials_isinventory_count"))
    eq_isinv = _int(isinv.get("equipment_isinventory_count"))
    if mat_isinv == 0 and eq_isinv == 0:
        return False
    setup = results.get("setup_data")
    if not setup or not isinstance(setup, dict):
        return False
    truck_with_tpl = _int(setup.get("truck_with_template"))
    wh_with_tpl = _int(setup.get("warehouse_with_template"))
    return truck_with_tpl > 0 or wh_with_tpl > 0


def _go_live_readiness(results: dict) -> tuple[bool, list[str]]:
    """Check if ready for go live: half of materials IsInventory + template checks. Returns (ready, list of what needs work)."""
    isinv = results.get("isinventory_counts") or {}
    setup = results.get("setup_data") or {}
    active_mats = _int(isinv.get("active_materials_count"))
    mat_isinv = _int(isinv.get("materials_isinventory_count"))
    truck_total = _int(setup.get("truck_total"))
    truck_with_tpl = _int(setup.get("truck_with_template"))
    wh_total = _int(setup.get("warehouse_total"))
    wh_with_tpl = _int(setup.get("warehouse_with_template"))
    templates_under_20 = _int(setup.get("templates_under_20_active_items"))

    needs: list[str] = []
    if active_mats > 0:
        pct = (100 * mat_isinv) / active_mats
        if pct < GO_LIVE_ITEMS_INVENTORY_MIN_PCT:
            needs.append(
                f"Mark at least half of pricebook materials as inventory (currently {mat_isinv} of {active_mats}, {pct:.0f}%)."
            )
    else:
        needs.append("Mark at least half of pricebook materials as inventory.")
    if truck_total > 0:
        truck_pct = (100 * truck_with_tpl) / truck_total
        if truck_pct < TRUCK_TEMPLATE_PCT_MIN:
            needs.append(
                f"Assign inventory templates to at least 80% of trucks (currently {truck_with_tpl} of {truck_total}, {truck_pct:.0f}%)."
            )
    if wh_total > 0 and wh_with_tpl < WAREHOUSE_WITH_TEMPLATE_MIN:
        needs.append("Assign an inventory template to at least one warehouse.")
    if templates_under_20 > 0:
        needs.append(
            f"Ensure all inventory templates have at least {TEMPLATE_MIN_ACTIVE_ITEMS} active items ({templates_under_20} template(s) have fewer)."
        )
    return (len(needs) == 0, needs)


def is_live_with_inventory(results: dict) -> bool:
    """True if inventory tracking start date (Inventory.Configuration BeginningDate) is today or in the past."""
    start = results.get("inventory_tracking_start_date")
    if not start:
        return False
    s = str(start).strip()[:10]  # "2024-09-10" or "2024-09-10T00:00:00"
    if len(s) < 10:
        return False
    try:
        track_start = datetime.strptime(s, "%Y-%m-%d").date()
        return track_start <= date.today()
    except (ValueError, TypeError):
        return False


def _int(v) -> int:
    """Coerce to int (e.g. from JSON string or Snowflake result)."""
    if v is None:
        return 0
    if isinstance(v, int):
        return v
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def load_results_from_combined_csv(path: Path, lookback_days: int = 90) -> dict:
    """Read combined audit query result from one CSV (header: source, v1..v16; 9 data rows)."""
    data = {"lookback_days": lookback_days}
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if len(rows) < 2:
        return data
    header = [str(x).strip().lower() if x else "" for x in rows[0]]
    if not header or header[0] != "source":
        return data
    col_index = {f"v{i}": i for i in range(1, 17)}
    for j, h in enumerate(header[1:], start=1):
        if h and h in col_index:
            col_index[h] = j
    known = {s.lower() for s in RESULT_KEYS}
    for r in rows[1:]:
        if not r:
            continue
        src = (r[0] or "").strip()
        if isinstance(src, (int, float)):
            src = str(int(src))
        src_lower = src.lower() if src else ""
        if src_lower in known:
            key = next(k for k in RESULT_KEYS if k.lower() == src_lower)
            values = [None] * 16
            for vi in range(1, 17):
                vkey = f"v{vi}"
                j = col_index.get(vkey, vi)
                if j < len(r):
                    values[vi - 1] = r[j]
            data[key] = [values]
    return data


def _excel_cell_value(cell) -> object:
    v = cell.value
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10] if hasattr(v, "date") else str(v)
    return v


def load_results_from_excel(path: Path, lookback_days: int = 90) -> dict:
    """Read combined audit result from one Excel sheet (source + v1..v16, 9 rows)."""
    try:
        import openpyxl
    except ImportError:
        print("Using --from-excel with .xlsx requires openpyxl. pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    data = {"lookback_days": lookback_days}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if len(wb.sheetnames) >= 1:
            ws = wb.active
            rows = [[_excel_cell_value(c) for c in row] for row in ws.iter_rows()]
            if len(rows) >= 2 and rows[0]:
                header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
                if header and header[0] == "source":
                    known = {s.lower() for s in RESULT_KEYS}
                    for r in rows[1:]:
                        if not r:
                            continue
                        src = (r[0] or "").strip()
                        if isinstance(src, (int, float)):
                            src = str(int(src))
                        if src.lower() in known:
                            key = next(k for k in RESULT_KEYS if k.lower() == src.lower())
                            values = list(r[1:17])
                            if len(values) < 16:
                                values = values + [None] * (16 - len(values))
                            data[key] = [values[:16]]
    finally:
        wb.close()
    return data


def load_results_from_csv_folder(folder: Path, lookback_days: int = 90) -> dict:
    """Read from CSVs: purchase_orders_summary.csv, invoice_materials.csv, replenishment_summary.csv, assessment_data.csv."""
    data = {"lookback_days": lookback_days}
    for key in RESULT_KEYS:
        path = folder / f"{key}.csv"
        if not path.is_file():
            continue
        rows = []
        try:
            with path.open(newline="", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                first = True
                for row in reader:
                    if first:
                        first = False
                        continue
                    rows.append(row)
        except Exception as e:
            print(f"Warning: could not read {path}: {e}", file=sys.stderr)
            continue
        if rows:
            data[key] = rows
    return data


def _snowflake_conn():
    """Create a Snowflake connection (caller must close). Requires SNOWFLAKE_ACCOUNT, SNOWFLAKE_USER."""
    try:
        import snowflake.connector
    except ImportError:
        print("Using --from-snowflake requires snowflake-connector-python. pip install snowflake-connector-python", file=sys.stderr)
        sys.exit(1)
    account = os.environ.get("SNOWFLAKE_ACCOUNT")
    user = os.environ.get("SNOWFLAKE_USER")
    if not account or not user:
        print("For --from-snowflake set SNOWFLAKE_ACCOUNT and SNOWFLAKE_USER in the environment.", file=sys.stderr)
        sys.exit(1)
    region = os.environ.get("SNOWFLAKE_REGION")
    if region and not account.endswith(f".{region}"):
        account = f"{account}.{region}"
    conn_params = {
        "account": account,
        "user": user,
        "authenticator": os.environ.get("SNOWFLAKE_AUTHENTICATOR", "externalbrowser"),
    }
    for key, env_key in (
        ("warehouse", "SNOWFLAKE_WAREHOUSE"),
        ("database", "SNOWFLAKE_DATABASE"),
        ("schema", "SNOWFLAKE_SCHEMA"),
        ("role", "SNOWFLAKE_ROLE"),
    ):
        val = os.environ.get(env_key)
        if val:
            conn_params[key] = val
    return snowflake.connector.connect(**conn_params)


def get_tenant_group_members(conn, group_name: str) -> list[tuple[int, str]]:
    """Return list of (tenant_id, tenant_name) for tenants in the given tenant group. Uses MASTER_DB."""
    query = """
    SELECT r._TENANT_ID, TRIM(COALESCE(r._TENANT_NAME, ''))
    FROM tenant_data.MASTER_DB.TENANTGROUP g
    JOIN tenant_data.MASTER_DB.TENANTRECORD_GROUPS_TENANTGROUP j ON g.ID = j.TENANTGROUP
    JOIN tenant_data.MASTER_DB.TENANTRECORD r ON j.TENANTRECORD = r._TENANT_ID
    WHERE g.NAME = %s
    ORDER BY r._TENANT_NAME
    """
    with conn.cursor() as cur:
        cur.execute(query, (group_name,))
        rows = cur.fetchall()
    return [(_int(row[0]), (row[1] or "").strip() or f"Tenant {row[0]}") for row in rows if row and row[0] is not None]


def load_results_from_snowflake(
    tenant_id: int,
    lookback_days: int = 90,
    sql_path: Path | None = None,
    conn=None,
) -> dict:
    """Run the combined audit query in Snowflake, return data dict for parse_results. If conn is provided, use it and do not close it."""
    sql_file = sql_path or (Path(__file__).resolve().parent / "sql" / "00_combined_audit.sql")
    if not sql_file.is_file():
        print(f"SQL file not found: {sql_file}", file=sys.stderr)
        sys.exit(1)
    raw_sql = sql_file.read_text(encoding="utf-8")
    if "SELECT 0 AS tenant_id" not in raw_sql:
        print("SQL file missing expected tenant_param placeholder (SELECT 0 AS tenant_id).", file=sys.stderr)
        sys.exit(1)
    sql = raw_sql.replace("SELECT 0 AS tenant_id", f"SELECT {int(tenant_id)} AS tenant_id")

    own_conn = False
    if conn is None:
        conn = _snowflake_conn()
        own_conn = True
    data = {"lookback_days": lookback_days}
    try:
        with conn.cursor() as cur:
            cur.execute(sql)
            rows = cur.fetchall()
        for row in rows:
            if not row or len(row) < 1:
                continue
            source = (row[0] or "").strip().lower()
            key = next((k for k in RESULT_KEYS if k.lower() == source), None)
            if key:
                values = list(row[1:17])
                if len(values) < 16:
                    values.extend([None] * (16 - len(values)))
                data[key] = [values[:16]]
    finally:
        if own_conn:
            conn.close()

    return data


def data_from_audit_rows(rows: list, lookback_days: int = 90) -> dict:
    """Build data dict for parse_results from audit query rows. Each row is [source, v1, v2, ..., v16] (e.g. from MCP or CSV)."""
    data = {"lookback_days": lookback_days}
    for row in rows:
        if not row or len(row) < 1:
            continue
        source = (row[0] or "").strip().lower()
        key = next((k for k in RESULT_KEYS if k.lower() == source), None)
        if key:
            values = list(row[1:17])
            if len(values) < 16:
                values.extend([None] * (16 - len(values)))
            data[key] = [values[:16]]
    return data


def parse_results(data: dict) -> dict:
    """Parse combined audit data (7 result sets) into a flat dict for evaluate_audit."""
    results = {}

    # Tenant info (for header): v1=tenant_id, v2=tenant_name
    ti = data.get("tenant_info")
    if ti is not None and isinstance(ti, list) and len(ti) >= 1:
        row = ti[0]
        if isinstance(row, (list, tuple)) and len(row) >= 2:
            results["tenant_id"] = row[0] if row[0] is not None and str(row[0]).strip() != "" else None
            results["tenant_name"] = (str(row[1]).strip() or None) if row[1] is not None else None
            if results["tenant_id"] is not None:
                results["tenant_id"] = _int(results["tenant_id"])

    # Pricebook (6 values): v1=material_count, v2=with_cost, v3=zero_cost, v4=with_vendor_link, v5=with_primary_vendor, v6=primary_default_replenishment
    pb = data.get("pricebook")
    if pb is not None and isinstance(pb, list) and len(pb) >= 1:
        row = pb[0]
        if isinstance(row, (list, tuple)) and len(row) >= 1:
            results["pricebook"] = {
                "material_count": _int(row[0]) if len(row) > 0 else 0,
                "materials_with_cost": _int(row[1]) if len(row) > 1 else 0,
                "materials_zero_cost": _int(row[2]) if len(row) > 2 else 0,
                "materials_with_vendor_link": _int(row[3]) if len(row) > 3 else 0,
                "materials_with_primary_vendor": _int(row[4]) if len(row) > 4 else 0,
                "primary_vendor_default_replenishment": _int(row[5]) if len(row) > 5 else 0,
            }
    if "pricebook" not in results:
        results["pricebook"] = {}

    # Purchase orders summary: v1=total_pending, v2=po_created_in_period, v3=earliest, v4=latest, v5..v11=status_0..6, v12=po_count, v13=single_line_pos, v14=total_lines, v15=placeholder_like, v16=pending_over_90
    pos = data.get("purchase_orders_summary")
    if pos is not None and isinstance(pos, list) and len(pos) >= 1:
        row = pos[0]
        if isinstance(row, (list, tuple)) and len(row) >= 11:
            results["total_pending_pos"] = _int(row[0])
            results["po_activity"] = {
                "po_created_in_period": _int(row[1]),
                "earliest_po": row[2] if len(row) > 2 else None,
                "latest_po": row[3] if len(row) > 3 else None,
            }
            by_status = {i: _int(row[4 + i]) for i in range(7) if 4 + i < len(row)}
            results["purchase_orders"] = {
                "by_status": by_status,
                "pending_count": sum(by_status.get(s, 0) for s in PENDING_STATUSES),
                "completed_count": sum(by_status.get(s, 0) for s in COMPLETED_STATUSES),
            }
            results["po_line_quality"] = {
                "po_count": _int(row[11]) if len(row) > 11 else 0,
                "pos_with_one_line_only": _int(row[12]) if len(row) > 12 else 0,
                "total_line_items": _int(row[13]) if len(row) > 13 else 0,
                "line_items_placeholder_like": _int(row[14]) if len(row) > 14 else 0,
            }
            results["pending_pos_over_90_days"] = _int(row[15]) if len(row) > 15 else 0
    if "total_pending_pos" not in results:
        results["total_pending_pos"] = 0
    if "po_activity" not in results:
        results["po_activity"] = {}
    if "purchase_orders" not in results:
        results["purchase_orders"] = {"by_status": {}, "pending_count": 0, "completed_count": 0}
    if "po_line_quality" not in results:
        results["po_line_quality"] = {}
    if "pending_pos_over_90_days" not in results:
        results["pending_pos_over_90_days"] = 0

    # Invoice materials: v1..v6 = line count, invoice count, tech-added, placeholder, invoices_gt_zero, invoices_gt_zero_with_material; v7..v10 = extended (IsInventory line count, zero-cost, from transfer, distinct materials used)
    inv = data.get("invoice_materials")
    if inv is not None and isinstance(inv, list) and len(inv) >= 1:
        row = inv[0]
        if isinstance(row, (list, tuple)):
            results["invoice_materials"] = {
                "material_line_count": _int(row[0]) if len(row) > 0 else 0,
                "invoices_with_materials": _int(row[1]) if len(row) > 1 else 0,
                "material_lines_added_by_technician": _int(row[2]) if len(row) > 2 else 0,
                "material_lines_placeholder_like": _int(row[3]) if len(row) > 3 else 0,
                "invoices_total_gt_zero": _int(row[4]) if len(row) > 4 else 0,
                "invoices_gt_zero_with_material": _int(row[5]) if len(row) > 5 else 0,
                "material_lines_IsInventory": _int(row[6]) if len(row) > 6 else 0,
                "material_lines_zero_cost": _int(row[7]) if len(row) > 7 else 0,
                "material_lines_from_transfer": _int(row[8]) if len(row) > 8 else 0,
                "distinct_IsInventory_materials_used_90d": _int(row[9]) if len(row) > 9 else 0,
            }
    if "invoice_materials" not in results:
        results["invoice_materials"] = {}

    # Replenishment: v1=open, v2=completed in lookback, v3=used_materials can be replenished
    repl = data.get("replenishment_summary")
    if repl is not None and isinstance(repl, list) and len(repl) >= 1:
        row = repl[0]
        if isinstance(row, (list, tuple)) and len(row) >= 2:
            results["replenishment_summary"] = {
                "open_count": _int(row[0]),
                "completed_in_lookback": _int(row[1]),
                "used_materials_can_be_replenished": _int(row[2]) if len(row) >= 3 else 0,
            }
    if "replenishment_summary" not in results:
        results["replenishment_summary"] = {"open_count": 0, "completed_in_lookback": 0, "used_materials_can_be_replenished": 0}

    # Returns summary: v1=total_pending, v2=in_period, v3=earliest, v4=latest, v5..v8=status_0..3, v9=pending_over_90_days
    rets = data.get("returns_summary")
    if rets is not None and isinstance(rets, list) and len(rets) >= 1:
        row = rets[0]
        if isinstance(row, (list, tuple)) and len(row) >= 5:
            results["total_pending_returns"] = _int(row[0])
            results["returns_activity"] = {
                "returns_in_period": _int(row[1]),
                "earliest_return": row[2] if len(row) > 2 else None,
                "latest_return": row[3] if len(row) > 3 else None,
            }
            results["returns_by_status"] = {i: _int(row[4 + i]) for i in range(4) if 4 + i < len(row)}
            if len(row) >= 9:
                results["pending_returns_over_90_days"] = _int(row[8])
    if "total_pending_returns" not in results:
        results["total_pending_returns"] = 0
    if "returns_by_status" not in results:
        results["returns_by_status"] = {}
    if "pending_returns_over_90_days" not in results:
        results["pending_returns_over_90_days"] = 0

    # Assessment: v1..v7 as before; v8..v13 = feature gates; v14 = materials with unit of measure
    ad = data.get("assessment_data")
    if ad is not None and isinstance(ad, list) and len(ad) >= 1:
        row = ad[0]
        if isinstance(row, (list, tuple)) and len(row) >= 5:
            results["readiness_config"] = {"purchasing_module_on": bool(_int(row[0])), "inventory_module_on": bool(_int(row[1]))}
            results["isinventory_counts"] = {
                "active_materials_count": _int(row[2]) if len(row) > 2 else 0,
                "active_equipment_count": _int(row[3]) if len(row) > 3 else 0,
                "materials_isinventory_count": _int(row[4]) if len(row) > 4 else 0,
                "equipment_isinventory_count": _int(row[5]) if len(row) > 5 else 0,
            }
            results["inventory_tracking_start_date"] = (str(row[6] or "").strip() or None) if len(row) > 6 else None
            if len(row) >= 15:
                results["feature_gates"] = {
                    "item_requisitions": bool(_int(row[7])),
                    "requisition_closeout": bool(_int(row[8])),
                    "install_requisitions": bool(_int(row[9])),
                    "transfers_to_jobs": bool(_int(row[10])),
                    "consignment": bool(_int(row[11])),
                    "purchase_approval_workflow": not bool(_int(row[12])),
                    "granular_wac": bool(_int(row[13])),
                }
                results["materials_with_unit_of_measure"] = _int(row[14])
            if len(row) >= 16:
                results["equipment_serialized_count"] = _int(row[15])
    if "readiness_config" not in results:
        results["readiness_config"] = {}
    if "isinventory_counts" not in results:
        results["isinventory_counts"] = {}
    if "inventory_tracking_start_date" not in results:
        results["inventory_tracking_start_date"] = None
    if "feature_gates" not in results:
        results["feature_gates"] = {}
    if "materials_with_unit_of_measure" not in results:
        results["materials_with_unit_of_measure"] = 0
    if "equipment_serialized_count" not in results:
        results["equipment_serialized_count"] = 0

    # Inventory settings: v1 = full JSON from NamedValue (Inventory.Configuration)
    inv_set = data.get("inventory_settings")
    if inv_set is not None and isinstance(inv_set, list) and len(inv_set) >= 1:
        row = inv_set[0]
        if isinstance(row, (list, tuple)) and len(row) > 0 and row[0]:
            raw = str(row[0]).strip()
            if raw:
                try:
                    results["inventory_settings"] = json.loads(raw)
                except json.JSONDecodeError:
                    results["inventory_settings"] = None
                else:
                    if not isinstance(results["inventory_settings"], dict):
                        results["inventory_settings"] = None
    if "inventory_settings" not in results:
        results["inventory_settings"] = None

    # Setup data: trucks, warehouses, templates with < 20 active items
    setup = data.get("setup_data")
    if setup is not None and isinstance(setup, list) and len(setup) >= 1:
        row = setup[0]
        if isinstance(row, (list, tuple)) and len(row) >= 5:
            results["setup_data"] = {
                "truck_total": _int(row[0]),
                "truck_with_template": _int(row[1]),
                "warehouse_total": _int(row[2]),
                "warehouse_with_template": _int(row[3]),
                "templates_under_20_active_items": _int(row[4]),
            }
    if "setup_data" not in results:
        results["setup_data"] = None

    # Usage checks: v1..v7 = repl_old_30d, transfers_old_14d, requisitions_old_90d, past_due_counts, negative_balances, direct_adj_90d, warehouses_no_count_90d; v8=completed_transfers_90d
    uc = data.get("usage_checks")
    if uc is not None and isinstance(uc, list) and len(uc) >= 1:
        row = uc[0]
        if isinstance(row, (list, tuple)) and len(row) >= 6:
            results["usage_checks"] = {
                "replenishment_older_than_30_days": _int(row[0]),
                "pending_transfers_older_than_14_days": _int(row[1]),
                "requisitions_older_than_90_days": _int(row[2]),
                "past_due_inventory_counts": _int(row[3]),
                "negative_balance_count": _int(row[4]),
                "direct_adjustments_in_90_days": _int(row[5]),
                "warehouses_no_completed_count_90d": _int(row[6]) if len(row) > 6 else 0,
                "completed_transfers_in_90_days": _int(row[7]) if len(row) > 7 else 0,
            }
    if "usage_checks" not in results:
        results["usage_checks"] = None

    return results


# Settings we actually care about (display order). Only these are shown; one line for tracking start.
SETTINGS_WE_CARE_ABOUT = [
    "beginningDate", "BeginningDate",
    "allowCopyingPoItemsToInvoice", "AllowCopyingPoItemsToInvoice",
    "allowNegativeQuantityOnHand", "AllowNegativeQuantityOnHand",
    "allowNegativeQuantityOnInvoice", "AllowNegativeQuantityOnInvoice",
    "autoApplyTagsToJobsBasedOnPoStatus", "AutoApplyTagsToJobsBasedOnPoStatus",
    "autoApplyTagsToJobsBasedOnTransferStatus", "AutoApplyTagsToJobsBasedOnTransferStatus",
    "autoAssignTruckToJobTransfers", "AutoAssignTruckToJobTransfers",
    "dontAutomaticallyCreateBills", "DontAutomaticallyCreateBills",
    "inventoryValuationMethod", "InventoryValuationMethod",
    "isBinTrackingEnabled", "IsBinTrackingEnabled",
    "isConsignmentInventoryTrackingEnabled", "IsConsignmentInventoryTrackingEnabled",
    "isInventoryMobileAppEnabled", "IsInventoryMobileAppEnabled",
    "isSerializedTrackingEnabled", "IsSerializedTrackingEnabled",
    "onlyReplenishMax", "OnlyReplenishMax",
    "poItemCostCopyOption", "PoItemCostCopyOption",
    "purchaseOrderApprovalField", "PurchaseOrderApprovalField",
]

# Human-readable labels for Inventory.Configuration keys (camelCase or PascalCase from app)
INVENTORY_SETTING_LABELS = {
    "beginningDate": "Inventory tracking start date",
    "BeginningDate": "Inventory tracking start date",
    "isBeginningDateStarted": "Tracking start date has passed",
    "IsBeginningDateStarted": "Tracking start date has passed",
    "allowNegativeQuantityOnHand": "Allow negative quantity on hand",
    "AllowNegativeQuantityOnHand": "Allow negative quantity on hand",
    "onlyReplenishMax": "Only replenish max",
    "OnlyReplenishMax": "Only replenish max",
    "isSerializedTrackingEnabled": "Serialized tracking enabled",
    "IsSerializedTrackingEnabled": "Serialized tracking enabled",
    "isSerialNumbersBulkUploadEnabled": "Serial numbers bulk upload enabled",
    "IsSerialNumbersBulkUploadEnabled": "Serial numbers bulk upload enabled",
    "isBinTrackingEnabled": "Bin tracking enabled",
    "IsBinTrackingEnabled": "Bin tracking enabled",
    "isSubAccountEnabled": "Sub-account enabled",
    "IsSubAccountEnabled": "Sub-account enabled",
    "trackingDisposition": "Tracking disposition",
    "TrackingDisposition": "Tracking disposition",
    "inventoryValuationMethod": "Inventory valuation method",
    "InventoryValuationMethod": "Inventory valuation method",
    "allowNegativeQuantityOnInvoice": "Allow negative quantity on invoice",
    "AllowNegativeQuantityOnInvoice": "Allow negative quantity on invoice",
    "dontAutomaticallyCreateBills": "Don't automatically create bills",
    "DontAutomaticallyCreateBills": "Don't automatically create bills",
    "hideCostOnMobilePurchaseOrders": "Hide cost on mobile POs",
    "HideCostOnMobilePurchaseOrders": "Hide cost on mobile POs",
    "allowCopyingPoItemsToInvoice": "Automatically add PO Items to Invoice",
    "AllowCopyingPoItemsToInvoice": "Automatically add PO Items to Invoice",
    "poItemCostCopyOption": "PO item cost copy option",
    "PoItemCostCopyOption": "PO item cost copy option",
    "printPoInLandscape": "Print PO in landscape",
    "PrintPoInLandscape": "Print PO in landscape",
    "printPoMinWidthLandscape": "Print PO min width (landscape)",
    "printPoMinWidthPortrait": "Print PO min width (portrait)",
    "showItemDescriptionWhenExportingPo": "Show item description when exporting PO",
    "ShowItemDescriptionWhenExportingPo": "Show item description when exporting PO",
    "autoApplyTagsToJobsBasedOnPoStatus": "Auto apply PO Tags to Jobs",
    "isInventoryMobileAppEnabled": "Inventory mobile app enabled",
    "IsInventoryMobileAppEnabled": "Inventory mobile app enabled",
    "displayServicesOnInvoiceCloseout": "Display services on invoice closeout",
    "DisplayServicesOnInvoiceCloseout": "Display services on invoice closeout",
    "isConsignmentInventoryTrackingEnabled": "Consignment inventory tracking enabled",
    "IsConsignmentInventoryTrackingEnabled": "Consignment inventory tracking enabled",
    "defaultConsignmentBusinessUnitId": "Default consignment business unit ID",
    "isMobileStorageEnabled": "Mobile storage enabled",
    "IsMobileStorageEnabled": "Mobile storage enabled",
    "autoAssignTruckToJobTransfers": "Assign Truck for Transfer (TTJ)",
    "sendUserNotificationsForPurchaseOrderReview": "Send user notifications for PO review",
    "sendApproverNotificationsForPurchaseOrderRequest": "Send approver notifications for PO request",
    "purchaseOrderApprovalField": "Purchase order approval field",
    "PurchaseOrderApprovalField": "Purchase order approval field",
    "requisitionDateNeededByType": "Requisition date needed by",
    "RequisitionDateNeededByType": "Requisition date needed by",
    "displayTaxesAndShippingOnPurchaseOrdersOnPdf": "Display taxes and shipping on PO PDF",
    "autoApplyTagsToJobsBasedOnTransferStatus": "Auto apply Transfer Tags to Jobs",
    "weightedAverageTrackByWarehouse": "Weighted average track by warehouse",
    "isWarehouseSitesEnabled": "Warehouse sites enabled",
    "IsWarehouseSitesEnabled": "Warehouse sites enabled",
    "maintainEstimateCostThroughRequisition": "Maintain estimate cost through requisition",
    "enableUnitOfMeasure": "Enable unit of measure",
    "EnableUnitOfMeasure": "Enable unit of measure",
}

INVENTORY_VALUATION_NAMES = {0: "Standard costing", 1: "Weighted average", 2: "Weighted average (granular)"}
TRACKING_DISPOSITION_NAMES = {0: "Singular", 1: "Dual"}
PO_APPROVAL_FIELD_NAMES = {0: "Purchase order total", 1: "Purchase order subtotal"}
REQUISITION_DATE_NEEDED_NAMES = {0: "30 days", 1: "1 week", 2: "2 weeks", 3: "90 days"}
PO_ITEM_COST_COPY_NAMES = {0: "Add items at $0", 1: "Add items at receipt cost"}


def format_inventory_settings(config: dict | None) -> list[str]:
    """Turn Inventory.Configuration JSON into readable lines. Only includes SETTINGS_WE_CARE_ABOUT; one line for tracking start."""
    if not config or not isinstance(config, dict):
        return []
    lines = []
    seen_keys_lower = set()
    for key in SETTINGS_WE_CARE_ABOUT:
        if key not in config:
            continue
        value = config[key]
        if value is None:
            continue
        if isinstance(value, (dict, list)):
            continue
        key_lower = (key or "").lower()
        if key_lower in seen_keys_lower:
            continue
        seen_keys_lower.add(key_lower)
        label = INVENTORY_SETTING_LABELS.get(key, key[:1].upper() + key[1:] if key else key)
        if isinstance(value, bool):
            lines.append(f"    {label}: {'Yes' if value else 'No'}")
        elif key_lower == "beginningdate" and value:
            lines.append(f"    {label}: {str(value)[:10]}")
        elif key_lower == "inventoryvaluationmethod":
            lines.append(f"    {label}: {INVENTORY_VALUATION_NAMES.get(value, value)}")
        elif key_lower == "trackingdisposition":
            lines.append(f"    {label}: {TRACKING_DISPOSITION_NAMES.get(value, value)}")
        elif key_lower == "purchaseorderapprovalfield":
            lines.append(f"    {label}: {PO_APPROVAL_FIELD_NAMES.get(value, value)}")
        elif key_lower == "requisitiondateneededbytype":
            lines.append(f"    {label}: {REQUISITION_DATE_NEEDED_NAMES.get(value, value)}")
        elif key_lower == "poitemcostcopyoption":
            lines.append(f"    {label}: {PO_ITEM_COST_COPY_NAMES.get(value, value)}")
        else:
            lines.append(f"    {label}: {value}")
    return lines


def _add(findings_by_area: dict, area: str, level: str, message: str) -> None:
    """Ensure area exists and append (level, message). level is 'green', 'yellow', 'review', or 'red'."""
    if area not in findings_by_area:
        findings_by_area[area] = {"green": [], "yellow": [], "review": [], "red": []}
    findings_by_area[area][level].append(message)


def _area_status(findings: dict) -> str:
    """Return one of four tiers from share of checks that are green: Red (Needs attention), Review, Yellow (Okay), Green (Good)."""
    g = len(findings.get("green") or [])
    y = len(findings.get("yellow") or [])
    rev = len(findings.get("review") or [])
    r = len(findings.get("red") or [])
    total = g + y + rev + r
    if total == 0:
        return "Green"
    pct_green = (100 * g) // total
    if pct_green >= 80:
        return "Green"   # Good
    if pct_green >= 60:
        return "Yellow"  # Okay (e.g. 3/5)
    if pct_green >= 40:
        return "Review"  # Needs review (e.g. 2/5)
    return "Red"         # Needs attention (0 or 1)


def get_action_plan(
    findings_by_area: dict[str, dict[str, list[str]]],
    audit_areas: list[str],
) -> list[str]:
    """Build a list of recommended actions from findings (non-Green areas). One or more actions per area as appropriate."""
    actions: list[str] = []
    seen: set[str] = set()  # dedupe identical recommendations

    def add(action: str) -> None:
        if action and action.strip() and action.strip() not in seen:
            seen.add(action.strip())
            actions.append(action.strip())

    for area in audit_areas:
        findings = findings_by_area.get(area, {"green": [], "yellow": [], "review": [], "red": []})
        status = _area_status(findings)
        if status == "Green":
            continue
        # Area needs attention; add relevant recommendation(s)
        if area == "Technician usage":
            add("Schedule technician training on adding materials to invoices in the field.")
        elif area == "Pricebook & setup":
            add("Review trucks and warehouses; assign inventory templates to trucks and warehouses as needed.")
            all_findings = (findings.get("red") or []) + (findings.get("review") or []) + (findings.get("yellow") or [])
            if any("20 active" in f or "fewer than" in f for f in all_findings):
                add("Review templates with fewer than 20 active items; add items or replace with a complete template.")
        elif area == "Purchasing":
            add("Review pending POs; receive or close old POs and keep pending count within half of truck count where possible.")
            add("Use multi-line POs for stock orders; avoid single-line or placeholder-heavy POs.")
        elif area == "Invoicing":
            add("Ensure invoices with amount > $0 include inventory materials where applicable; train office and field on invoice material usage.")
            all_findings = (findings.get("red") or []) + (findings.get("review") or [])
            if any("Zero-cost" in f or "zero-cost" in f for f in all_findings):
                add("Review cost flow from PO/receiving to invoice; fix zero-cost IsInventory lines where cost should flow.")
        elif area == "Replenishment":
            add("Address open replenishment requests; complete and receive replenishments to improve turnover.")
        elif area == "Returns":
            add("Process pending returns and issue credits so returns don't age.")
        elif area == "Transfers":
            add("Complete or cancel aging transfers; encourage transfer-to-job usage where appropriate.")
        elif area == "Counts and adjustments":
            all_findings = (findings.get("red") or []) + (findings.get("review") or []) + (findings.get("yellow") or [])
            count_added = False
            if any("Past due" in f or "past due" in f for f in all_findings):
                add("Complete or reschedule past due inventory counts.")
                count_added = True
            if any("Negative" in f or "negative" in f for f in all_findings):
                add("Run an inventory count and adjustments to resolve negative quantities.")
                count_added = True
            if any("Direct" in f and "adjustment" in f for f in all_findings):
                add("Review direct adjustment practices; reduce manual adjustments where possible.")
                count_added = True
            if any("no completed count" in f or "no count" in f for f in all_findings):
                add("Schedule and complete cycle counts for all warehouses.")
                count_added = True
            if not count_added:
                add("Review counts and adjustments; complete cycle counts and resolve discrepancies as needed.")
        elif area == "Other":
            add("Close or complete aging requisitions.")

    return actions


def evaluate_audit(results: dict, lookback_days: int) -> tuple[bool, dict[str, dict[str, list[str]]]]:
    """Run audit checks (POs, invoices, replenishment, etc.). Returns (ok, findings_by_area). ok = no red in any area."""
    findings_by_area: dict[str, dict[str, list[str]]] = {}

    # --- Purchasing ---
    po = results.get("purchase_orders", {})
    by_status = po.get("by_status", {})
    pending_waiting = by_status.get(PENDING_STATUS_WAITING_TO_SEND, 0)
    completed = po.get("completed_count", 0)
    po_activity = results.get("po_activity", {})
    total_pending = results.get("total_pending_pos", 0)
    pending_po_over_90 = results.get("pending_pos_over_90_days", 0)
    po_created_in_lookback = po_activity.get("po_created_in_period", 0)
    plq = results.get("po_line_quality", {})
    total_in_lookback = sum(by_status.values()) if by_status else 0
    po_count_in_lookback = plq.get("po_count", 0)
    denom = po_created_in_lookback if po_created_in_lookback > 0 else (total_in_lookback if total_in_lookback > 0 else po_count_in_lookback)

    area_po = "Purchasing"
    setup_po = results.get("setup_data")
    truck_total_po = (setup_po.get("truck_total", 0) or 0) if setup_po and isinstance(setup_po, dict) else 0
    allowed_pending = max(0, int(truck_total_po * PENDING_POS_ALLOWED_PER_TRUCK)) if truck_total_po > 0 else None

    msg_po = f"Total pending POs: {total_pending}."
    if pending_po_over_90 > 0:
        msg_po += f" {pending_po_over_90} of these are older than 90 days."
    if allowed_pending is not None and allowed_pending > 0:
        pending_pct_of_allowed = (100 * total_pending) // allowed_pending
        msg_po += f" Allowed (half of {truck_total_po} trucks): {allowed_pending}; pending is {pending_pct_of_allowed}% of allowed."
        if pending_pct_of_allowed <= PENDING_PCT_OF_ALLOWED_GOOD_MAX:
            _add(findings_by_area, area_po, "green", msg_po)
        elif pending_pct_of_allowed <= PENDING_PCT_OF_ALLOWED_OKAY_MAX:
            _add(findings_by_area, area_po, "yellow", msg_po)
        elif pending_pct_of_allowed <= PENDING_PCT_OF_ALLOWED_REVIEW_MAX:
            _add(findings_by_area, area_po, "review", msg_po)
        else:
            _add(findings_by_area, area_po, "red", msg_po)
    else:
        if total_pending > TOTAL_PENDING_ABSOLUTE_RED or pending_po_over_90 > 0:
            _add(findings_by_area, area_po, "red", msg_po)
        else:
            _add(findings_by_area, area_po, "green", msg_po)

    if allowed_pending is not None and allowed_pending > 0 and (total_pending > 0 or pending_po_over_90 > 0):
        over_90_pct = (100 * pending_po_over_90) // allowed_pending
        msg_90 = f"Pending POs over 90 days: {pending_po_over_90} ({over_90_pct}% of allowed pending {allowed_pending})."
        if over_90_pct <= PENDING_OVER_90_PCT_GOOD_MAX:
            _add(findings_by_area, area_po, "green", msg_90)
        elif over_90_pct <= PENDING_OVER_90_PCT_OKAY_MAX:
            _add(findings_by_area, area_po, "yellow", msg_90)
        elif over_90_pct <= PENDING_OVER_90_PCT_REVIEW_MAX:
            _add(findings_by_area, area_po, "review", msg_90)
        else:
            _add(findings_by_area, area_po, "red", msg_90)

    if denom == 0:
        _add(findings_by_area, area_po, "red", "No purchase order activity in the lookback period. Purchasing module may not be in use.")
    else:
        if completed > 0 or po_created_in_lookback > 0:
            _add(findings_by_area, area_po, "green", f"POs created in last {lookback_days} days: {po_created_in_lookback}; completed (received/exported): {completed}.")
        if po_created_in_lookback > 0:
            receive_rate_pct = (100 * completed) // po_created_in_lookback
            msg_rc = f"PO receive rate: {receive_rate_pct}% of POs created in period are received/exported (target ≥70%)."
            if receive_rate_pct >= PO_RECEIVE_RATE_GOOD_MIN_PCT:
                status_rc = "green"
            elif receive_rate_pct >= PO_RECEIVE_RATE_OKAY_MIN_PCT:
                status_rc = "yellow"
            elif receive_rate_pct >= PO_RECEIVE_RATE_REVIEW_MIN_PCT:
                status_rc = "review"
            else:
                status_rc = "red"
            _add(findings_by_area, area_po, status_rc, msg_rc)
        if pending_waiting > 5 and completed == 0:
            _add(findings_by_area, area_po, "red", f"Many pending POs ({pending_waiting}) but no completed POs. Are receipts being created?")

    plq_po_count = plq.get("po_count", 0)
    plq_single = plq.get("pos_with_one_line_only", 0)
    plq_total_lines = plq.get("total_line_items", 0)
    plq_placeholder = plq.get("line_items_placeholder_like", 0)
    if plq_po_count > 0:
        multi = plq_po_count - plq_single
        multi_pct = (100 * multi) // plq_po_count
        line_msg = f"PO line items: {plq_total_lines} lines across {plq_po_count} POs; {multi} of {plq_po_count} ({multi_pct}%) have multiple lines (<40% Needs attention, 40-50% Needs review, 50-60% Okay, >60% Good)."
        if multi_pct < PO_MULTI_LINE_RED_MAX_PCT:
            _add(findings_by_area, area_po, "red", line_msg)
        elif multi_pct < PO_MULTI_LINE_REVIEW_MAX_PCT:
            _add(findings_by_area, area_po, "review", line_msg)
        elif PO_MULTI_LINE_YELLOW_MIN_PCT <= multi_pct <= PO_MULTI_LINE_YELLOW_MAX_PCT:
            _add(findings_by_area, area_po, "yellow", line_msg)
        else:
            _add(findings_by_area, area_po, "green", line_msg)
        if plq_total_lines > 0 and (plq_placeholder / plq_total_lines) > PO_PLACEHOLDER_MAJORITY:
            _add(findings_by_area, area_po, "red", f"Majority of PO line items mention placeholder/generic/material ({plq_placeholder} of {plq_total_lines}). Use specific parts.")
        elif plq_total_lines > 0 and plq_placeholder > 0:
            _add(findings_by_area, area_po, "green", f"PO line descriptions: {plq_placeholder} of {plq_total_lines} lines with placeholder-like text (under 50%).")
        elif plq_total_lines > 0:
            _add(findings_by_area, area_po, "green", "PO line descriptions: majority are specific parts.")

    # --- Invoicing ---
    inv = results.get("invoice_materials", {})
    inv_total_gt_zero = inv.get("invoices_total_gt_zero", 0)
    inv_gt_zero_with_material = inv.get("invoices_gt_zero_with_material", 0)
    area_inv = "Invoicing"
    if inv_total_gt_zero == 0:
        _add(findings_by_area, area_inv, "yellow", "No invoices with amount greater than zero in the lookback period; invoice material coverage not assessed.")
    else:
        pct_with_material = (100 * inv_gt_zero_with_material) // inv_total_gt_zero
        line = f"Of {inv_total_gt_zero} invoice(s) with amount > $0, {inv_gt_zero_with_material} ({pct_with_material}%) have at least one IsInventory material line (85%+ Good, 75-85% Okay, 50-75% Needs review, <50% Needs attention)."
        if pct_with_material >= INVOICE_GT_ZERO_WITH_MATERIAL_GOOD_MIN_PCT:
            _add(findings_by_area, area_inv, "green", line)
        elif pct_with_material >= INVOICE_GT_ZERO_WITH_MATERIAL_OKAY_MIN_PCT:
            _add(findings_by_area, area_inv, "yellow", line)
        elif pct_with_material >= INVOICE_GT_ZERO_WITH_MATERIAL_REVIEW_MIN_PCT:
            _add(findings_by_area, area_inv, "review", line)
        else:
            _add(findings_by_area, area_inv, "red", line)

    # Extended: % of material lines that are IsInventory
    mat_total = inv.get("material_line_count", 0)
    mat_isinv = inv.get("material_lines_IsInventory", 0)
    if mat_total > 0 and mat_isinv is not None:
        pct_isinv = (100 * mat_isinv) // mat_total
        line2 = f"Of {mat_total} invoice material lines, {mat_isinv} ({pct_isinv}%) are IsInventory (85%+ Good, 75-85% Okay, 50-75% Needs review, <50% Needs attention)."
        if pct_isinv >= INVOICE_MATERIAL_LINES_ISINVENTORY_GOOD_MIN_PCT:
            s2 = "green"
        elif pct_isinv >= INVOICE_MATERIAL_LINES_ISINVENTORY_OKAY_MIN_PCT:
            s2 = "yellow"
        elif pct_isinv >= INVOICE_MATERIAL_LINES_ISINVENTORY_REVIEW_MIN_PCT:
            s2 = "review"
        else:
            s2 = "red"
        _add(findings_by_area, area_inv, s2, line2)
    # Extended: zero-cost IsInventory lines (< 5% green, 5–15% yellow, >= 15% red)
    if mat_isinv and mat_isinv > 0:
        zero_cost = inv.get("material_lines_zero_cost", 0)
        pct_zero = (100 * zero_cost) // mat_isinv
        if pct_zero < INVOICE_ZERO_COST_GOOD_MAX_PCT:
            _add(findings_by_area, area_inv, "green", f"Zero-cost IsInventory lines: {zero_cost} of {mat_isinv} ({pct_zero}%).")
        elif pct_zero < INVOICE_ZERO_COST_OKAY_MAX_PCT:
            _add(findings_by_area, area_inv, "yellow", f"Zero-cost IsInventory lines: {zero_cost} of {mat_isinv} ({pct_zero}%).")
        elif pct_zero < INVOICE_ZERO_COST_REVIEW_MAX_PCT:
            _add(findings_by_area, area_inv, "review", f"Zero-cost IsInventory lines: {zero_cost} of {mat_isinv} ({pct_zero}%). Review cost flow if unexpected.")
        else:
            _add(findings_by_area, area_inv, "red", f"Zero-cost IsInventory lines: {zero_cost} of {mat_isinv} ({pct_zero}%). Review cost flow from PO/receiving.")

    # --- Technician usage ---
    mat_lines = inv.get("material_line_count", 0)
    added_by_tech = inv.get("material_lines_added_by_technician", 0)
    area_tech = "Technician usage"
    if mat_lines > 0:
        pct_tech = (100 * added_by_tech) // mat_lines
        tech_line = f"{added_by_tech} of {mat_lines} invoice material line(s) ({pct_tech}%) were added by a technician (≥60% Good, 50-60% Okay, 30-50% Needs review, <30% Needs attention)."
        if pct_tech >= INVOICE_TECH_ADDED_GOOD_MIN_PCT:
            _add(findings_by_area, area_tech, "green", tech_line)
        elif pct_tech >= INVOICE_TECH_ADDED_OKAY_MIN_PCT:
            _add(findings_by_area, area_tech, "yellow", tech_line)
        elif pct_tech >= INVOICE_TECH_ADDED_REVIEW_MIN_PCT:
            _add(findings_by_area, area_tech, "review", tech_line)
        else:
            _add(findings_by_area, area_tech, "red", tech_line)
    else:
        _add(findings_by_area, area_tech, "red", "0 of 0 invoice material lines were added by a technician (no material line activity; 0%).")

    # --- Replenishment ---
    repl = results.get("replenishment_summary", {})
    repl_open = repl.get("open_count", 0)
    repl_completed = repl.get("completed_in_lookback", 0)
    repl_msg = f"Replenishment: {repl_open} open (pending/in progress), {repl_completed} completed in last {lookback_days} days."
    area_repl = "Replenishment"
    if repl_completed > 0:
        rate_pct = 100 * (1 - (repl_open / repl_completed))
        if rate_pct >= REPLENISHMENT_RATE_GOOD_MIN_PCT:
            _add(findings_by_area, area_repl, "green", repl_msg)
        elif rate_pct >= REPLENISHMENT_RATE_OKAY_MIN_PCT:
            _add(findings_by_area, area_repl, "yellow", repl_msg)
        elif rate_pct >= REPLENISHMENT_RATE_REVIEW_MIN_PCT:
            _add(findings_by_area, area_repl, "review", repl_msg)
        else:
            _add(findings_by_area, area_repl, "red", repl_msg)
    elif repl_open > 0:
        _add(findings_by_area, area_repl, "red", repl_msg + " No completed replenishments to compare; address open requests.")
    else:
        _add(findings_by_area, area_repl, "red", repl_msg + " No replenishment activity in lookback (0 of 0); treat as 0%.")

    uc = results.get("usage_checks")
    if uc and isinstance(uc, dict):
        repl_30 = uc.get("replenishment_older_than_30_days", 0)
        if repl_30 > USAGE_REPLENISHMENT_OLD_30D_MAX:
            _add(findings_by_area, area_repl, "red", f"Replenishment requests older than 30 days: {repl_30} (max {USAGE_REPLENISHMENT_OLD_30D_MAX}). Address aging replenishment requests.")

    # --- Returns ---
    ret_by_status = results.get("returns_by_status", {})
    ret_received = ret_by_status.get(RETURN_CREDIT_RECEIVED_STATUS, 0)
    total_pending_ret = results.get("total_pending_returns", 0)
    pending_over_90 = results.get("pending_returns_over_90_days", 0)
    ret_activity = results.get("returns_activity") or {}
    returns_in_period = ret_activity.get("returns_in_period", 0)
    ret_msg = f"Returns: {ret_received} credit received of {returns_in_period} returns in last {lookback_days} days. {total_pending_ret} pending."
    if returns_in_period > 0 and total_pending_ret > 0:
        pct_pending = (100 * total_pending_ret) // returns_in_period
        ret_msg += f" ({pct_pending}% of returns in period)"
    if pending_over_90 > 0:
        ret_msg += f" {pending_over_90} pending are older than 90 days."
    area_ret = "Returns"
    if pending_over_90 > 0:
        _add(findings_by_area, area_ret, "yellow", ret_msg + " Returns to be resolved in timely manner.")
    elif returns_in_period > 0 and total_pending_ret > 0:
        pct_pending = (100 * total_pending_ret) // returns_in_period
        if pct_pending <= RETURN_PENDING_PCT_GOOD_MAX:
            _add(findings_by_area, area_ret, "green", ret_msg)
        elif pct_pending <= RETURN_PENDING_PCT_OKAY_MAX:
            _add(findings_by_area, area_ret, "yellow", ret_msg)
        elif pct_pending <= RETURN_PENDING_PCT_REVIEW_MAX:
            _add(findings_by_area, area_ret, "review", ret_msg + " Returns to be resolved in timely manner.")
        else:
            _add(findings_by_area, area_ret, "red", ret_msg + " Returns to be resolved in timely manner.")
    elif total_pending_ret > 0:
        _add(findings_by_area, area_ret, "yellow", ret_msg + " Returns to be resolved in timely manner.")
    elif returns_in_period == 0:
        _add(findings_by_area, area_ret, "red", ret_msg + " No returns activity in lookback (0 of 0); treat as 0%.")
    else:
        _add(findings_by_area, area_ret, "green", ret_msg)

    # --- Pricebook & setup ---
    setup = results.get("setup_data")
    area_setup = "Pricebook & setup"
    if setup and isinstance(setup, dict):
        wh_total = setup.get("warehouse_total", 0)
        wh_with_tpl = setup.get("warehouse_with_template", 0)
        templates_under = setup.get("templates_under_20_active_items", 0)
        truck_total = setup.get("truck_total", 0)
        truck_with_tpl = setup.get("truck_with_template", 0)
        truck_pct = (100 * truck_with_tpl // truck_total) if truck_total > 0 else 0
        truck_ok = truck_total == 0 or truck_pct >= TRUCK_TEMPLATE_PCT_MIN
        if truck_total > 0:
            if truck_ok:
                _add(findings_by_area, area_setup, "green", f"Trucks: {truck_with_tpl} of {truck_total} with inventory template (≥80%).")
            else:
                _add(findings_by_area, area_setup, "red", f"Trucks: {truck_with_tpl} of {truck_total} with inventory template (min 80% expected).")
        if wh_total > 0 and wh_with_tpl < WAREHOUSE_WITH_TEMPLATE_MIN:
            _add(findings_by_area, area_setup, "red", f"Warehouses with inventory template: {wh_with_tpl} of {wh_total}. At least {WAREHOUSE_WITH_TEMPLATE_MIN} warehouse(s) must have a template assigned.")
        elif wh_total > 0:
            _add(findings_by_area, area_setup, "green", f"Warehouses: {wh_with_tpl} of {wh_total} with inventory template.")
        if templates_under > 0:
            _add(findings_by_area, area_setup, "red", f"Templates with fewer than {TEMPLATE_MIN_ACTIVE_ITEMS} active items: {templates_under}. Each template should have at least {TEMPLATE_MIN_ACTIVE_ITEMS} active items.")
        elif setup.get("truck_total", 0) + wh_total > 0:
            _add(findings_by_area, area_setup, "green", f"All templates have at least {TEMPLATE_MIN_ACTIVE_ITEMS} active items.")

    # (Materials with no unit of measure: informational only in setup summary; not scored.)
    # Extended: IsInventory materials with no invoice usage in 90 days
    mat_isinv_total = (results.get("isinventory_counts") or {}).get("materials_isinventory_count", 0)
    distinct_used = inv.get("distinct_IsInventory_materials_used_90d", 0)
    if mat_isinv_total > 0 and distinct_used is not None:
        unused = mat_isinv_total - distinct_used
        if unused > 0:
            pct_unused = (100 * unused) // mat_isinv_total
            un_line = f"IsInventory materials with no invoice usage in last {lookback_days} days: {unused} of {mat_isinv_total} ({pct_unused}%)."
            sun = "red" if pct_unused >= MATERIALS_UNUSED_90D_RED_MIN_PCT else "yellow" if pct_unused > MATERIALS_UNUSED_90D_YELLOW_MAX_PCT else "green"
            _add(findings_by_area, area_setup, sun, un_line)

    # --- Transfers ---
    area_transfers = "Transfers"
    if uc and isinstance(uc, dict):
        trans_14 = uc.get("pending_transfers_older_than_14_days", 0)
        completed_transfers = uc.get("completed_transfers_in_90_days", 0)
        truck_total = (setup or {}).get("truck_total", 0)
        # Expect half of trucks to get at least one transfer per week over lookback
        expected_min_transfers = max(0, int(truck_total * TRANSFER_EXPECTED_TRUCK_FRACTION * (lookback_days / 7))) if truck_total else 0
        if trans_14 > USAGE_TRANSFERS_OLD_14D_MAX:
            _add(findings_by_area, area_transfers, "red", f"Pending transfers older than 14 days: {trans_14} (max {USAGE_TRANSFERS_OLD_14D_MAX}). Complete or cancel aging transfers.")
        else:
            _add(findings_by_area, area_transfers, "green", f"Pending transfers older than 14 days: {trans_14} (max {USAGE_TRANSFERS_OLD_14D_MAX}).")
        if truck_total > 0:
            if completed_transfers >= expected_min_transfers:
                _add(findings_by_area, area_transfers, "green", f"Completed transfers in last {lookback_days} days: {completed_transfers} (expected ≥{expected_min_transfers} for ~{int(TRANSFER_EXPECTED_TRUCK_FRACTION * 100)}% of {truck_total} trucks at least once per week).")
            else:
                _add(findings_by_area, area_transfers, "red", f"Completed transfers in last {lookback_days} days: {completed_transfers} (expected ≥{expected_min_transfers} for ~{int(TRANSFER_EXPECTED_TRUCK_FRACTION * 100)}% of {truck_total} trucks at least once per week). Increase transfer usage.")

    # --- Counts and adjustments ---
    area_counts = "Counts and adjustments"
    if uc and isinstance(uc, dict):
        past_due = uc.get("past_due_inventory_counts", 0)
        neg_bal = uc.get("negative_balance_count", 0)
        direct_adj = uc.get("direct_adjustments_in_90_days", 0)
        wh_no_count = uc.get("warehouses_no_completed_count_90d", 0)
        if past_due > USAGE_PAST_DUE_COUNTS_MAX:
            _add(findings_by_area, area_counts, "red", f"Past due inventory counts: {past_due} (max {USAGE_PAST_DUE_COUNTS_MAX}). Complete or reschedule counts.")
        else:
            _add(findings_by_area, area_counts, "green", f"Past due inventory counts: {past_due} (max {USAGE_PAST_DUE_COUNTS_MAX}).")
        if neg_bal > 0:
            _add(findings_by_area, area_counts, "red", f"Negative inventory balances: {neg_bal} location/SKU combination(s) with negative quantity. Resolve negative balances.")
        if direct_adj > USAGE_DIRECT_ADJUSTMENTS_MAX:
            _add(findings_by_area, area_counts, "red", f"Direct (quantity) adjustments in last 90 days: {direct_adj} (max {USAGE_DIRECT_ADJUSTMENTS_MAX}). High volume may indicate process or data issues.")
        else:
            _add(findings_by_area, area_counts, "green", f"Direct (quantity) adjustments in last 90 days: {direct_adj} (max {USAGE_DIRECT_ADJUSTMENTS_MAX}).")
        if wh_no_count > WAREHOUSES_NO_COUNT_90D_MAX:
            _add(findings_by_area, area_counts, "red", f"Warehouses with no completed count in last 90 days: {wh_no_count}. Schedule and complete cycle counts.")
        elif (setup or {}).get("warehouse_total", 0) > 0:
            _add(findings_by_area, area_counts, "green", "All warehouses have had a completed count in the last 90 days.")

    # --- Other (requisitions: not required to use, but bad if many open/aging) ---
    area_other = "Other"
    if uc and isinstance(uc, dict):
        req_90 = uc.get("requisitions_older_than_90_days", 0)
        if req_90 >= USAGE_REQUISITIONS_OLD_90D_MAX:
            _add(findings_by_area, area_other, "red", f"Requisitions older than 90 days: {req_90} (max {USAGE_REQUISITIONS_OLD_90D_MAX}). Resolve or close aging requisitions.")
        else:
            _add(findings_by_area, area_other, "green", f"Requisitions older than 90 days: {req_90} (max {USAGE_REQUISITIONS_OLD_90D_MAX}).")

    ok = not any(f.get("red") for f in findings_by_area.values())
    return ok, findings_by_area


def evaluate_audit_purchasing_only(results: dict, lookback_days: int) -> tuple[bool, dict[str, dict[str, list[str]]], bool]:
    """Purchasing-only audit when not yet live with inventory. Includes Pricebook, Purchasing, Invoicing, Replenishment, Returns.
    Returns (ok, findings_by_area, ready_for_inventory_implementation). ready = no red in Pricebook, Purchasing, or Invoicing."""
    findings_by_area: dict[str, dict[str, list[str]]] = {}

    # --- Pricebook (from Inventory_Prepardness_Check: zero cost %, default replenishment vendor %) ---
    area_pb = "Pricebook"
    pb = results.get("pricebook", {})
    mat_count = pb.get("material_count", 0)
    zero_cost = pb.get("materials_zero_cost", 0)
    def_repl = pb.get("primary_vendor_default_replenishment", 0)

    if mat_count == 0:
        _add(findings_by_area, area_pb, "red", "Pricebook has no materials. Add materials before starting inventory implementation.")
    else:
        zero_pct = (zero_cost / mat_count) if mat_count else 0
        if zero_cost > 0 and zero_pct > PRICEBOOK_RED_PCT:
            _add(findings_by_area, area_pb, "red", f"{zero_cost} material(s) have $0 cost ({zero_pct:.0%} of {mat_count}). Assign costs before inventory (red when >{PRICEBOOK_RED_PCT:.0%}).")
        elif zero_cost > 0:
            _add(findings_by_area, area_pb, "green", f"{zero_cost} material(s) have $0 cost ({zero_pct:.1%} of {mat_count}); under {PRICEBOOK_RED_PCT:.0%} threshold.")
        else:
            _add(findings_by_area, area_pb, "green", "All materials have a cost assigned.")
        def_repl_pct = (def_repl / mat_count) if mat_count else 0
        if def_repl > 0 and def_repl_pct > PRICEBOOK_RED_PCT:
            _add(findings_by_area, area_pb, "red", f"{def_repl} material(s) have Default or Imported Default Replenishment Vendor as primary ({def_repl_pct:.0%} of {mat_count}). Assign a real primary vendor (red when >{PRICEBOOK_RED_PCT:.0%}).")
        elif def_repl > 0:
            _add(findings_by_area, area_pb, "green", f"{def_repl} material(s) use Default/Imported Default Replenishment as primary ({def_repl_pct:.1%}); under threshold.")
        else:
            _add(findings_by_area, area_pb, "green", "No materials use Default or Imported Default Replenishment Vendor as primary.")

    # --- Purchasing (same as full audit) ---
    po = results.get("purchase_orders", {})
    by_status = po.get("by_status", {})
    pending_waiting = by_status.get(PENDING_STATUS_WAITING_TO_SEND, 0)
    completed = po.get("completed_count", 0)
    po_activity = results.get("po_activity", {})
    total_pending = results.get("total_pending_pos", 0)
    pending_po_over_90 = results.get("pending_pos_over_90_days", 0)
    po_created_in_lookback = po_activity.get("po_created_in_period", 0)
    plq = results.get("po_line_quality", {})
    total_in_lookback = sum(by_status.values()) if by_status else 0
    po_count_in_lookback = plq.get("po_count", 0)
    denom = po_created_in_lookback if po_created_in_lookback > 0 else (total_in_lookback if total_in_lookback > 0 else po_count_in_lookback)

    area_po = "Purchasing"
    setup_po = results.get("setup_data")
    truck_total_po = (setup_po.get("truck_total", 0) or 0) if setup_po and isinstance(setup_po, dict) else 0
    allowed_pending = max(0, int(truck_total_po * PENDING_POS_ALLOWED_PER_TRUCK)) if truck_total_po > 0 else None

    msg_po = f"Total pending POs: {total_pending}."
    if pending_po_over_90 > 0:
        msg_po += f" {pending_po_over_90} of these are older than 90 days."
    if allowed_pending is not None and allowed_pending > 0:
        pending_pct_of_allowed = (100 * total_pending) // allowed_pending
        msg_po += f" Allowed (half of {truck_total_po} trucks): {allowed_pending}; pending is {pending_pct_of_allowed}% of allowed."
        if pending_pct_of_allowed <= PENDING_PCT_OF_ALLOWED_GOOD_MAX:
            _add(findings_by_area, area_po, "green", msg_po)
        elif pending_pct_of_allowed <= PENDING_PCT_OF_ALLOWED_OKAY_MAX:
            _add(findings_by_area, area_po, "yellow", msg_po)
        elif pending_pct_of_allowed <= PENDING_PCT_OF_ALLOWED_REVIEW_MAX:
            _add(findings_by_area, area_po, "review", msg_po)
        else:
            _add(findings_by_area, area_po, "red", msg_po)
    else:
        if total_pending > TOTAL_PENDING_ABSOLUTE_RED or pending_po_over_90 > 0:
            _add(findings_by_area, area_po, "red", msg_po)
        else:
            _add(findings_by_area, area_po, "green", msg_po)

    if allowed_pending is not None and allowed_pending > 0 and (total_pending > 0 or pending_po_over_90 > 0):
        over_90_pct = (100 * pending_po_over_90) // allowed_pending
        msg_90 = f"Pending POs over 90 days: {pending_po_over_90} ({over_90_pct}% of allowed pending {allowed_pending})."
        if over_90_pct <= PENDING_OVER_90_PCT_GOOD_MAX:
            _add(findings_by_area, area_po, "green", msg_90)
        elif over_90_pct <= PENDING_OVER_90_PCT_OKAY_MAX:
            _add(findings_by_area, area_po, "yellow", msg_90)
        elif over_90_pct <= PENDING_OVER_90_PCT_REVIEW_MAX:
            _add(findings_by_area, area_po, "review", msg_90)
        else:
            _add(findings_by_area, area_po, "red", msg_90)

    if denom == 0:
        _add(findings_by_area, area_po, "red", "No purchase order activity in the lookback period. Purchasing module may not be in use.")
    else:
        if completed > 0 or po_created_in_lookback > 0:
            _add(findings_by_area, area_po, "green", f"POs created in last {lookback_days} days: {po_created_in_lookback}; completed (received/exported): {completed}.")
        if po_created_in_lookback > 0:
            receive_rate_pct = (100 * completed) // po_created_in_lookback
            msg_rc = f"PO receive rate: {receive_rate_pct}% of POs created in period are received/exported (target ≥70%)."
            if receive_rate_pct >= PO_RECEIVE_RATE_GOOD_MIN_PCT:
                status_rc = "green"
            elif receive_rate_pct >= PO_RECEIVE_RATE_OKAY_MIN_PCT:
                status_rc = "yellow"
            elif receive_rate_pct >= PO_RECEIVE_RATE_REVIEW_MIN_PCT:
                status_rc = "review"
            else:
                status_rc = "red"
            _add(findings_by_area, area_po, status_rc, msg_rc)
        if pending_waiting > 5 and completed == 0:
            _add(findings_by_area, area_po, "red", f"Many pending POs ({pending_waiting}) but no completed POs. Are receipts being created?")

    plq_po_count = plq.get("po_count", 0)
    plq_single = plq.get("pos_with_one_line_only", 0)
    plq_total_lines = plq.get("total_line_items", 0)
    plq_placeholder = plq.get("line_items_placeholder_like", 0)
    if plq_po_count > 0:
        multi = plq_po_count - plq_single
        multi_pct = (100 * multi) // plq_po_count
        line_msg = f"PO line items: {plq_total_lines} lines across {plq_po_count} POs; {multi} of {plq_po_count} ({multi_pct}%) have multiple lines (<40% Needs attention, 40-50% Needs review, 50-60% Okay, >60% Good)."
        if multi_pct < PO_MULTI_LINE_RED_MAX_PCT:
            _add(findings_by_area, area_po, "red", line_msg)
        elif multi_pct < PO_MULTI_LINE_REVIEW_MAX_PCT:
            _add(findings_by_area, area_po, "review", line_msg)
        elif PO_MULTI_LINE_YELLOW_MIN_PCT <= multi_pct <= PO_MULTI_LINE_YELLOW_MAX_PCT:
            _add(findings_by_area, area_po, "yellow", line_msg)
        else:
            _add(findings_by_area, area_po, "green", line_msg)
        if plq_total_lines > 0 and (plq_placeholder / plq_total_lines) > PO_PLACEHOLDER_MAJORITY:
            _add(findings_by_area, area_po, "red", f"Majority of PO line items mention placeholder/generic/material ({plq_placeholder} of {plq_total_lines}). Use specific parts.")
        elif plq_total_lines > 0 and plq_placeholder > 0:
            _add(findings_by_area, area_po, "green", f"PO line descriptions: {plq_placeholder} of {plq_total_lines} lines with placeholder-like text (under 50%).")
        elif plq_total_lines > 0:
            _add(findings_by_area, area_po, "green", "PO line descriptions: majority are specific parts.")

    # --- Replenishment (same as full audit) ---
    repl = results.get("replenishment_summary", {})
    repl_open = repl.get("open_count", 0)
    repl_completed = repl.get("completed_in_lookback", 0)
    repl_msg = f"Replenishment: {repl_open} open (pending/in progress), {repl_completed} completed in last {lookback_days} days."
    area_repl = "Replenishment"
    if repl_completed > 0:
        rate_pct = 100 * (1 - (repl_open / repl_completed))
        if rate_pct >= REPLENISHMENT_RATE_GOOD_MIN_PCT:
            _add(findings_by_area, area_repl, "green", repl_msg)
        elif rate_pct >= REPLENISHMENT_RATE_OKAY_MIN_PCT:
            _add(findings_by_area, area_repl, "yellow", repl_msg)
        elif rate_pct >= REPLENISHMENT_RATE_REVIEW_MIN_PCT:
            _add(findings_by_area, area_repl, "review", repl_msg)
        else:
            _add(findings_by_area, area_repl, "red", repl_msg)
    elif repl_open > 0:
        _add(findings_by_area, area_repl, "red", repl_msg + " No completed replenishments to compare; address open requests.")
    else:
        _add(findings_by_area, area_repl, "red", repl_msg + " No replenishment activity in lookback (0 of 0); treat as 0%.")

    uc = results.get("usage_checks")
    if uc and isinstance(uc, dict):
        repl_30 = uc.get("replenishment_older_than_30_days", 0)
        if repl_30 > USAGE_REPLENISHMENT_OLD_30D_MAX:
            _add(findings_by_area, area_repl, "red", f"Replenishment requests older than 30 days: {repl_30} (max {USAGE_REPLENISHMENT_OLD_30D_MAX}). Address aging replenishment requests.")

    # --- Returns (same as full audit) ---
    ret_by_status = results.get("returns_by_status", {})
    ret_received = ret_by_status.get(RETURN_CREDIT_RECEIVED_STATUS, 0)
    total_pending_ret = results.get("total_pending_returns", 0)
    pending_over_90 = results.get("pending_returns_over_90_days", 0)
    ret_activity = results.get("returns_activity") or {}
    returns_in_period = ret_activity.get("returns_in_period", 0)
    ret_msg = f"Returns: {ret_received} credit received of {returns_in_period} returns in last {lookback_days} days. {total_pending_ret} pending."
    if returns_in_period > 0 and total_pending_ret > 0:
        pct_pending = (100 * total_pending_ret) // returns_in_period
        ret_msg += f" ({pct_pending}% of returns in period)"
    if pending_over_90 > 0:
        ret_msg += f" {pending_over_90} pending are older than 90 days."
    area_ret = "Returns"
    if pending_over_90 > 0:
        _add(findings_by_area, area_ret, "yellow", ret_msg + " Returns to be resolved in timely manner.")
    elif returns_in_period > 0 and total_pending_ret > 0:
        pct_pending = (100 * total_pending_ret) // returns_in_period
        if pct_pending <= RETURN_PENDING_PCT_GOOD_MAX:
            _add(findings_by_area, area_ret, "green", ret_msg)
        elif pct_pending <= RETURN_PENDING_PCT_OKAY_MAX:
            _add(findings_by_area, area_ret, "yellow", ret_msg)
        elif pct_pending <= RETURN_PENDING_PCT_REVIEW_MAX:
            _add(findings_by_area, area_ret, "review", ret_msg + " Returns to be resolved in timely manner.")
        else:
            _add(findings_by_area, area_ret, "red", ret_msg + " Returns to be resolved in timely manner.")
    elif total_pending_ret > 0:
        _add(findings_by_area, area_ret, "yellow", ret_msg + " Returns to be resolved in timely manner.")
    elif returns_in_period == 0:
        _add(findings_by_area, area_ret, "red", ret_msg + " No returns activity in lookback (0 of 0); treat as 0%.")
    else:
        _add(findings_by_area, area_ret, "green", ret_msg)

    # --- Invoicing (no IsInventory-specific checks: no % with IsInventory, no zero-cost IsInventory, no from-transfer) ---
    inv = results.get("invoice_materials", {})
    inv_total_gt_zero = inv.get("invoices_total_gt_zero", 0)
    mat_lines = inv.get("material_line_count", 0)
    added_by_tech = inv.get("material_lines_added_by_technician", 0)
    area_inv = "Invoicing"
    if inv_total_gt_zero == 0:
        _add(findings_by_area, area_inv, "yellow", "No invoices with amount greater than zero in the lookback period.")
    else:
        _add(findings_by_area, area_inv, "green", f"Invoices with amount > $0 in last {lookback_days} days: {inv_total_gt_zero}; material lines: {mat_lines}.")
    if mat_lines > 0:
        pct_tech = (100 * added_by_tech) // mat_lines
        tech_line = f"{added_by_tech} of {mat_lines} invoice material line(s) ({pct_tech}%) were added by a technician (≥60% Good, 50-60% Okay, 30-50% Needs review, <30% Needs attention)."
        if pct_tech >= INVOICE_TECH_ADDED_GOOD_MIN_PCT:
            _add(findings_by_area, area_inv, "green", tech_line)
        elif pct_tech >= INVOICE_TECH_ADDED_OKAY_MIN_PCT:
            _add(findings_by_area, area_inv, "yellow", tech_line)
        elif pct_tech >= INVOICE_TECH_ADDED_REVIEW_MIN_PCT:
            _add(findings_by_area, area_inv, "review", tech_line)
        else:
            _add(findings_by_area, area_inv, "red", tech_line)
    elif inv_total_gt_zero > 0:
        _add(findings_by_area, area_inv, "red", "0 of 0 invoice material lines were added by a technician (no material line activity; 0%).")

    # Ready for inventory implementation: pricebook good, purchasing good, invoicing good (no red in those three)
    area_po = "Purchasing"
    area_inv = "Invoicing"
    pb_ok = not (findings_by_area.get(area_pb, {}).get("red"))
    po_ok = not (findings_by_area.get(area_po, {}).get("red"))
    inv_ok = not (findings_by_area.get(area_inv, {}).get("red"))
    ready_for_inventory_implementation = bool(pb_ok and po_ok and inv_ok)

    ok = not any(f.get("red") for f in findings_by_area.values())
    return ok, findings_by_area, ready_for_inventory_implementation


def _escape(s: str) -> str:
    """Escape for HTML text content."""
    return html.escape(str(s), quote=True)


# Short labels for group summary table (same as scorecard badges)
# Four tiers: Needs attention (0–1 green checks), Needs review (2), Okay (3), Good (4–5)
STATUS_LABEL = {"Red": "Critical", "Review": "Review", "Yellow": "On track", "Green": "Good", "N/A": "N/A"}


def _group_overall_score(tenant_results: list[dict]) -> tuple[int, int, int]:
    """Return (good_cells, total_cells, pct). Each cell = one tenant × one area; N/A excluded."""
    good = 0
    total = 0
    for t in tenant_results:
        area_statuses = t.get("area_statuses") or {}
        for area in AUDIT_AREAS:
            status = area_statuses.get(area, "Green")
            if status == "N/A":
                continue
            total += 1
            if status == "Green":
                good += 1
    pct = (100 * good // total) if total else 0
    return good, total, pct


def _group_common_struggles(tenant_results: list[dict], min_tenants: int = 2) -> list[tuple[str, int, int]]:
    """Return list of (area, count_struggling, total_with_area) for areas where at least min_tenants struggle (not Good). Sorted by count_struggling desc."""
    by_area: dict[str, list[int]] = {a: [] for a in AUDIT_AREAS}  # per area: 1 = struggling (not Good), 0 = Good, -1 = N/A
    for t in tenant_results:
        area_statuses = t.get("area_statuses") or {}
        for area in AUDIT_AREAS:
            status = area_statuses.get(area, "Green")
            if status == "N/A":
                by_area[area].append(-1)
            elif status == "Green":
                by_area[area].append(0)
            else:
                by_area[area].append(1)
    out: list[tuple[str, int, int]] = []
    for area in AUDIT_AREAS:
        vals = by_area[area]
        total_with = sum(1 for v in vals if v >= 0)  # not N/A
        struggling = sum(1 for v in vals if v == 1)
        if total_with > 0 and struggling >= min_tenants:
            out.append((area, struggling, total_with))
    out.sort(key=lambda x: (-x[1], x[0]))
    return out


def _group_global_action_plan(tenant_results: list[dict], min_tenants: int = 2) -> list[tuple[str, int]]:
    """Return list of (action, count) for actions that appear in at least min_tenants, sorted by count desc."""
    from collections import Counter
    counts: Counter[str] = Counter()
    for t in tenant_results:
        plan = t.get("action_plan") or []
        for a in plan:
            if a and isinstance(a, str) and a.strip():
                counts[a.strip()] += 1
    return [(a, c) for a, c in counts.most_common() if c >= min_tenants]


def render_group_summary_html(
    group_name: str,
    tenant_results: list[dict],
    lookback: int,
) -> str:
    """Build group summary HTML: header, overall score, common struggles, global action plan, table of tenants with area statuses and links to scorecard_<id>.html."""
    title = f"Inventory Scorecard — {_escape(group_name)} (group summary)"
    n = len(tenant_results)
    n_need_attention = sum(1 for t in tenant_results if not t.get("ok", True))
    summary_line = (
        f"{n_need_attention} of {n} tenant(s) need attention in at least one area."
        if n_need_attention > 0
        else f"All {n} tenants are in good shape (no red areas)."
    )

    # Group score for summary (no "X of Y" detail)
    good_cells, total_cells, score_pct = _group_overall_score(tenant_results)
    if total_cells > 0:
        score_label = "Strong" if score_pct >= 80 else "On track" if score_pct >= 60 else "Needs focus" if score_pct >= 40 else "Critical"
        score_class = "ok" if score_pct >= 60 else "not-ok"
        summary_score_line = f'<p class="overall {score_class}">Group health: {score_pct}% — {_escape(score_label)}</p>'
    else:
        summary_score_line = ""

    # Common struggles (areas where at least 2 tenants are not Good)
    common = _group_common_struggles(tenant_results, min_tenants=2)
    if common:
        common_items = "".join(
            f'<li>{_escape(area)}: {count} of {total} tenant(s) need improvement</li>'
            for area, count, total in common
        )
        common_struggles_html = f'<ul class="group-common-list">\n        {common_items}\n      </ul>'
    else:
        common_struggles_html = "<p>No single area stands out as a common struggle across multiple tenants.</p>"

    # Global action plan (actions that appear in at least 2 tenants)
    global_actions = _group_global_action_plan(tenant_results, min_tenants=2)
    if global_actions:
        global_items = "".join(
            f'<li>{_escape(action)} <span class="global-action-count">({count} tenant{"s" if count != 1 else ""})</span></li>'
            for action, count in global_actions
        )
        global_action_html = f'<ol class="group-action-plan-list">\n        {global_items}\n      </ol>'
    else:
        global_action_html = "<p>No actions apply to multiple tenants; review individual scorecards for tenant-specific recommendations.</p>"

    # Table header
    area_headers = "".join(
        f'<th scope="col">{_escape(a)}</th>' for a in AUDIT_AREAS
    )
    rows_html = []
    for t in tenant_results:
        tid = t.get("tenant_id") or 0
        tname = _escape(t.get("tenant_name") or f"Tenant {tid}")
        link = f'<a href="scorecard_{tid}.html">{tname}</a>'
        area_statuses = t.get("area_statuses") or {}
        cells = []
        for area in AUDIT_AREAS:
            status = area_statuses.get(area, "Green")
            sl = "na" if status == "N/A" else status.lower()
            label = STATUS_LABEL.get(status, status)
            cells.append(f'<td class="status-{sl}"><span class="badge badge-{sl}">{_escape(label)}</span></td>')
        rows_html.append(
            f"        <tr><td class=\"tenant-name\">{link}</td>\n            "
            + "\n            ".join(cells)
            + "\n        </tr>"
        )

    html_doc = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{title}</title>
  <link rel="stylesheet" href="scorecard.css">
  <style>
    .group-summary-wrap {{ overflow-x: auto; width: 100%; margin-top: 0.5rem; }}
    .group-summary-table {{ width: 100%; min-width: 600px; border-collapse: collapse; }}
    .group-summary-table th, .group-summary-table td {{ padding: 0.4rem 0.5rem; text-align: left; border-bottom: 1px solid var(--st-border, #dfe0e1); font-size: 0.8125rem; }}
    .group-summary-table th {{ font-weight: 600; color: var(--st-text-secondary, #3f3f46); white-space: nowrap; }}
    .group-summary-table .tenant-name {{ font-weight: 500; }}
    .group-summary-table a {{ color: var(--st-primary, #2270ee); text-decoration: none; }}
    .group-summary-table a:hover {{ text-decoration: underline; }}
    .group-summary-table td.status-red {{ background: var(--danger-muted, #fef2f2); }}
    .group-summary-table td.status-review {{ background: #fff7ed; }}
    .group-summary-table td.status-yellow {{ background: var(--warn-muted, #fffbeb); }}
    .group-summary-table td.status-green {{ background: var(--success-muted, #ecfdf5); }}
    .group-summary-table td.status-na {{ background: var(--st-surface-alt, #fafafa); color: var(--st-text-muted, #71717a); }}
    .group-summary-table .badge-review {{ background: #fed7aa; color: #9a3412; }}
    .group-summary-table .badge-na {{ background: #e5e7eb; color: #6b7280; }}
    .group-subhead {{ font-size: 0.875rem; font-weight: 600; margin: 1rem 0 0.35rem 0; color: var(--st-text-secondary, #3f3f46); }}
    .group-subhead:first-of-type {{ margin-top: 0; }}
    .group-common-list {{ margin: 0; padding-left: 1.25rem; font-size: 0.9375rem; }}
    .group-common-list li {{ margin: 0.35rem 0; }}
    .group-action-plan-list {{ margin: 0; padding-left: 1.35rem; font-size: 0.9375rem; }}
    .group-action-plan-list li {{ margin: 0.4rem 0; }}
    .global-action-count {{ color: var(--st-text-muted, #71717a); font-size: 0.8125rem; font-weight: 500; }}
    .card-desc {{ margin: 0 0 0.5rem 0; font-size: 0.8125rem; color: var(--st-text-muted, #71717a); }}
  </style>
</head>
<body>
  <div class="page">
    <header class="header">
      <h1>Inventory Setup &amp; Usage Scorecard — Group Summary</h1>
      <p class="meta">
        <span>{_escape(group_name)}</span>
        <span>Last {lookback} days</span>
        <span>{n} tenant(s)</span>
      </p>
    </header>

    <div class="card scorecard">
      <h2 class="card-title">Summary</h2>
      <p class="overall {'ok' if n_need_attention == 0 else 'not-ok'}">{_escape(summary_line)}</p>
      {summary_score_line}
      <p>Click a tenant name to open their full scorecard.</p>
    </div>

    <section class="card block">
      <h2 class="card-title">Common struggles &amp; action plan</h2>
      <h3 class="group-subhead">Common struggles</h3>
      <p class="card-desc">Areas where multiple tenants need improvement.</p>
      {common_struggles_html}
      <h3 class="group-subhead">Recommended actions</h3>
      <p class="card-desc">Actions that apply to multiple tenants in this group.</p>
      {global_action_html}
    </section>

    <section class="card block">
      <h2 class="card-title">Tenants by area</h2>
      <div class="group-summary-wrap">
      <table class="group-summary-table" aria-label="Tenant scorecard status by area">
        <thead>
          <tr>
            <th scope="col">Tenant</th>
            {area_headers}
          </tr>
        </thead>
        <tbody>
{chr(10).join(rows_html)}
        </tbody>
      </table>
      </div>
    </section>
  </div>
</body>
</html>
"""
    return html_doc


def render_scorecard_html(
    tenant_id: int | None,
    tenant_name: str | None,
    lookback: int,
    results: dict,
    findings_by_area: dict[str, dict[str, list[str]]],
    ok: bool,
    audit_areas: list[str] | None = None,
    is_live: bool = True,
    ready_for_inventory_implementation: bool | None = None,
) -> str:
    """Build a self-contained, customer-friendly HTML scorecard. All dynamic text is escaped. When is_live=False, title is Purchasing Audit and only audit_areas are shown. When not is_live and ready_for_inventory_implementation is True, show verdict 'Ready to start inventory implementation'."""
    audit_areas = audit_areas or AUDIT_AREAS
    tenant_display = _escape(tenant_name or f"Tenant {tenant_id}" if tenant_id else "Inventory Audit")
    if is_live:
        title = f"Inventory Setup &amp; Usage Scorecard — {tenant_display}"
    else:
        title = f"Purchasing Audit Scorecard — {tenant_display}"

    status_class = {"Red": "status-red", "Review": "status-review", "Yellow": "status-yellow", "Green": "status-green"}
    status_label = {"Red": "Critical", "Review": "Review", "Yellow": "On track", "Green": "Good"}

    # Scorecard grid HTML (for external CSS layout)
    scorecard_items = []
    for area in audit_areas:
        findings = findings_by_area.get(area, {"green": [], "yellow": [], "review": [], "red": []})
        status = _area_status(findings)
        status_lower = status.lower()  # red, review, yellow, green
        label = status_label.get(status, status)
        scorecard_items.append(
            f'        <div class="scorecard-item status-{status_lower}">'
            f'<span class="area-name">{_escape(area)}</span>'
            f'<span class="badge badge-{status_lower}">{_escape(label)}</span></div>'
        )

    # Findings by area HTML
    findings_sections = []
    for area in audit_areas:
        findings = findings_by_area.get(area, {"green": [], "yellow": [], "review": [], "red": []})
        if not any(findings.values()):
            continue
        status = _area_status(findings)
        label = status_label.get(status, status)
        lines = []
        for g in findings["green"]:
            lines.append(f'          <li class="finding-ok">{_escape(g)}</li>')
        for y in findings["yellow"]:
            lines.append(f'          <li class="finding-warn">{_escape(y)}</li>')
        for rev in findings.get("review") or []:
            lines.append(f'          <li class="finding-review">{_escape(rev)}</li>')
        for r in findings["red"]:
            lines.append(f'          <li class="finding-err">{_escape(r)}</li>')
        border_cls = "border-green" if status == "Green" else "border-yellow" if status == "Yellow" else "border-review" if status == "Review" else "border-red"
        findings_sections.append(
            f'      <section class="findings-area {border_cls}">\n'
            f'        <h3 class="area-head">{_escape(area)} <span class="badge badge-{status.lower()}">{_escape(label)}</span></h3>\n'
            f'        <ul class="findings-list">\n' + "\n".join(lines) + "\n"
            f"        </ul>\n      </section>"
        )

    # Settings lines (when not is_live: only feature gates + tracking start date; no inventory-specific UI settings)
    cfg = results.get("readiness_config") or {}
    fg = results.get("feature_gates") or {}
    settings_lines = [
        f"<p><strong>Purchasing module:</strong> {'On' if cfg.get('purchasing_module_on') else 'Off' if cfg.get('purchasing_module_on') is False else '—'}",
        f"<p><strong>Inventory module:</strong> {'On' if cfg.get('inventory_module_on') else 'Off' if cfg.get('inventory_module_on') is False else '—'}",
    ]
    if fg:
        settings_lines.append(f"<p><strong>Item requisitions:</strong> {'On' if fg.get('item_requisitions') else 'Off'}")
        settings_lines.append(f"<p><strong>Requisition closeout:</strong> {'On' if fg.get('requisition_closeout') else 'Off'}")
        settings_lines.append(f"<p><strong>Transfers to jobs:</strong> {'On' if fg.get('transfers_to_jobs') else 'Off'}")
        settings_lines.append(f"<p><strong>Consignment:</strong> {'On' if fg.get('consignment') else 'Off'}")
        settings_lines.append(f"<p><strong>Purchase approval workflow:</strong> {'On' if fg.get('purchase_approval_workflow') else 'Off'}")
        settings_lines.append(f"<p><strong>Granular WAC:</strong> {'On' if fg.get('granular_wac') else 'Off'}")
    inv_settings = results.get("inventory_settings")
    if inv_settings:
        if is_live:
            for line in format_inventory_settings(inv_settings):
                # line is "    Label: value"; strip and escape
                parts = line.strip().split(":", 1)
                if len(parts) == 2:
                    settings_lines.append(f"<p><strong>{_escape(parts[0].strip())}:</strong> {_escape(parts[1].strip())}")
                else:
                    settings_lines.append(f"<p>{_escape(line.strip())}")
        else:
            # Purchasing audit: show only tracking start date for context
            start = inv_settings.get("BeginningDate") or inv_settings.get("beginningDate")
            if start is not None:
                start_str = str(start).strip()[:10]
                settings_lines.append(f"<p><strong>Inventory tracking start date:</strong> {_escape(start_str)}</p>")
    settings_html = "\n    ".join(settings_lines) if settings_lines else "<p>Not available.</p>"

    # Setup summary (when not is_live: omit entirely — trucks/warehouses/templates are inventory-only)
    setup_parts = []
    if is_live:
        isinv = results.get("isinventory_counts") or {}
        setup = results.get("setup_data")
        active_mats = isinv.get("active_materials_count")
        mat_isinv = isinv.get("materials_isinventory_count")
        uom_count = results.get("materials_with_unit_of_measure", 0)
        active_eq = isinv.get("active_equipment_count")
        eq_isinv = isinv.get("equipment_isinventory_count")
        eq_serialized = results.get("equipment_serialized_count", 0)
        setup_parts = [
            f"<p><strong>Materials:</strong> total {active_mats if active_mats is not None else '—'}; "
            f"with IsInventory: {mat_isinv if mat_isinv is not None else '—'}; with unit of measure: {uom_count}.",
            f"<p><strong>Equipment:</strong> total {active_eq if active_eq is not None else '—'}; "
            f"with IsInventory: {eq_isinv if eq_isinv is not None else '—'}; serialized: {eq_serialized}.",
        ]
        if setup and isinstance(setup, dict):
            truck_total = setup.get("truck_total", 0)
            truck_with_tpl = setup.get("truck_with_template", 0)
            wh_total = setup.get("warehouse_total", 0)
            wh_with_tpl = setup.get("warehouse_with_template", 0)
            templates_under = setup.get("templates_under_20_active_items", 0)
            truck_pct = (100 * truck_with_tpl // truck_total) if truck_total > 0 else 0
            truck_ok = truck_total == 0 or truck_pct >= TRUCK_TEMPLATE_PCT_MIN
            wh_ok = wh_total == 0 or wh_with_tpl >= WAREHOUSE_WITH_TEMPLATE_MIN
            templates_ok = templates_under == 0
            setup_parts.append(
                f"<p><strong>Trucks:</strong> {truck_total} total, {truck_with_tpl} with inventory template "
                f"(min 80% expected) — {'OK' if truck_ok else 'Needs attention'}."
            )
            setup_parts.append(
                f"<p><strong>Warehouses:</strong> {wh_total} total, {wh_with_tpl} with template — {'OK' if wh_ok else 'Needs attention'}."
            )
            setup_parts.append(
                f"<p><strong>Templates with &lt; {TEMPLATE_MIN_ACTIVE_ITEMS} active items:</strong> {templates_under} — {'OK' if templates_ok else 'Needs attention'}."
            )
    setup_html = "\n    ".join(setup_parts) if setup_parts else "<p>Not available.</p>"
    setup_section_html = (
        "    <section class=\"card block\">\n      <h2 class=\"card-title\">Setup summary</h2>\n      "
        + setup_html
        + "\n    </section>\n\n    "
    ) if is_live else ""

    overall_msg = "No critical issues. Review any yellow areas as needed." if ok else "Address red areas to improve setup and usage."
    # Readiness box (purchasing audit only): Go Live Readiness if preparing, else Inventory Readiness
    readiness_section_html = ""
    if not is_live:
        if _is_preparing_for_go_live(results, is_live):
            go_live_ready, go_live_needs = _go_live_readiness(results)
            if go_live_ready:
                readiness_body = "<p class=\"overall ok\">Ready to set your go live date and do inventory beginning balance counts.</p>"
            else:
                needs_list = "\n      ".join(f"<li>{_escape(n)}</li>" for n in go_live_needs)
                readiness_body = f"<ul class=\"readiness-needs\">\n      {needs_list}\n    </ul>"
            readiness_section_html = (
                "    <section class=\"card block go-live-readiness\">\n"
                "      <h2 class=\"card-title\">Go Live Readiness</h2>\n      "
                + readiness_body + "\n"
                "    </section>\n\n    "
            )
        elif ready_for_inventory_implementation is not None:
            if ready_for_inventory_implementation is True:
                readiness_body = "<p class=\"overall ok\">Ready to start inventory implementation.</p>"
            else:
                areas_needing_work = _readiness_areas_needing_work(findings_by_area)
                areas_str = _format_areas_for_message(areas_needing_work)
                if areas_str:
                    readiness_body = f"<p class=\"overall not-ok\">Address red items in {_escape(areas_str)} to prepare for inventory implementation.</p>"
                else:
                    readiness_body = "<p class=\"overall not-ok\">Address red items above to prepare for inventory implementation.</p>"
            readiness_section_html = (
                "    <section class=\"card block inventory-readiness\">\n"
                "      <h2 class=\"card-title\">Inventory Readiness</h2>\n      "
                + readiness_body + "\n"
                "    </section>\n\n    "
            )

    # Action plan from findings (non-Green areas)
    action_plan = get_action_plan(findings_by_area, audit_areas)
    if action_plan:
        action_plan_html = "<ol class=\"action-plan-list\">\n      " + "\n      ".join(
            f"<li>{_escape(a)}</li>" for a in action_plan
        ) + "\n    </ol>"
    else:
        action_plan_html = "<p>No specific actions recommended; keep up current practices.</p>"
    action_plan_section = (
        "    <section class=\"card block action-plan\">\n"
        "      <h2 class=\"card-title\">Recommended actions</h2>\n      "
        + action_plan_html
        + "\n    </section>\n\n    "
    )

    h1_text = "Inventory Setup &amp; Usage Scorecard" if is_live else "Purchasing Audit Scorecard"
    meta_extra = ""
    if not is_live:
        start = results.get("inventory_tracking_start_date")
        start_str = str(start).strip()[:10] if start else "—"
        meta_extra = f'<span>Not yet live with inventory (tracking start: {_escape(start_str)})</span>'

    html_doc = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{_escape(title)}</title>
  <link rel="stylesheet" href="scorecard.css">
</head>
<body>
  <div class="page">
    <header class="header">
      <h1>{h1_text}</h1>
      <p class="meta">
        <span>{tenant_display}</span>
        <span>Last {lookback} days</span>
        {f'<span>Tenant ID: {tenant_id}</span>' if tenant_id is not None else ''}
        {meta_extra}
      </p>
    </header>

    <div class="card scorecard">
      <h2 class="card-title">Scorecard by area</h2>
      <div class="scorecard-grid">
{chr(10).join(scorecard_items)}
      </div>
      <p class="overall {'ok' if ok else 'not-ok'}">{_escape(overall_msg)}</p>
    </div>

    {readiness_section_html}    <section class="card block">
      <h2 class="card-title">Settings</h2>
      {settings_html}
    </section>

    {setup_section_html}    <section class="card block">
      <h2 class="card-title">Findings by area</h2>
{chr(10).join(findings_sections)}
    </section>

    {action_plan_section}  </div>
</body>
</html>
"""
    return html_doc


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Inventory Setup/Usage Audit: report on POs, invoices, and replenishment for tenants on the inventory module."
    )
    parser.add_argument("--tenant-id", type=int, default=None, help="Tenant ID for report header")
    parser.add_argument("--tenant-name", type=str, default=None, help="Tenant name for report header")
    parser.add_argument("--lookback-days", type=int, default=90, help="Lookback days (default 90)")
    parser.add_argument("--from-excel", type=str, default=None, metavar="FILE", help="Read from Excel (.xlsx) or CSV (combined 4-row export)")
    parser.add_argument("--from-csv-folder", type=str, default=None, metavar="FOLDER", help="Read from CSVs in a folder (purchase_orders_summary.csv, etc.)")
    parser.add_argument("--from-results", type=str, default=None, metavar="JSON_FILE", help="Read from a JSON file")
    parser.add_argument("--from-group-json", type=str, default=None, metavar="JSON_FILE", help="Read group audit results from JSON (see TENANT_GROUP_DISPLAY.md); requires --html for group summary path")
    parser.add_argument("--from-snowflake", action="store_true", help="Run the audit query in Snowflake (requires --tenant-id or --tenant-group; uses Okta via SNOWFLAKE_* env)")
    parser.add_argument("--tenant-group", type=str, default=None, metavar="NAME", help="Run audit for all tenants in this tenant group (requires --from-snowflake and --html)")
    parser.add_argument("--html", type=str, default=None, metavar="FILE", help="Write a customer-friendly HTML scorecard to FILE (for --tenant-group, write group summary here and scorecard_<id>.html alongside)")
    parser.add_argument("--output-aggregate", type=str, default=None, metavar="FILE", help="Merge this run into an aggregated JSON file (for hosted lookup by tenant). Use with single or group run.")
    args = parser.parse_args()

    data = None
    lookback = args.lookback_days

    # --- Group JSON mode: load pre-fetched audit rows (e.g. from MCP) and write group summary + per-tenant scorecards ---
    if args.from_group_json and args.html:
        path = Path(args.from_group_json)
        if not path.is_file():
            print(f"Group JSON file not found: {path}", file=sys.stderr)
            return 1
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError as e:
            print(f"Invalid JSON in {path}: {e}", file=sys.stderr)
            return 1
        group_name = raw.get("group_name") or "Tenant group"
        lookback = _int(raw.get("lookback_days")) or args.lookback_days
        tenants_in = raw.get("tenants") or []
        if not isinstance(tenants_in, list):
            print("Group JSON must have 'tenants' array.", file=sys.stderr)
            return 1
        base = Path(args.html).resolve().parent
        output_dir = base / _sanitize_folder_name(group_name)
        output_dir.mkdir(parents=True, exist_ok=True)
        css_src = Path(__file__).resolve().parent / "css" / "scorecard.css"
        if css_src.is_file():
            (output_dir / "scorecard.css").write_text(css_src.read_text(encoding="utf-8"), encoding="utf-8")
        tenant_results = []
        for t_in in tenants_in:
            tid = _int(t_in.get("tenant_id"))
            tname = (t_in.get("tenant_name") or "").strip() or f"Tenant {tid}"
            rows = t_in.get("rows") or []
            if not rows:
                print(f"  Skipping {tname} (no rows).", file=sys.stderr)
                continue
            data = data_from_audit_rows(rows, lookback)
            results = parse_results(data)
            is_live = is_live_with_inventory(results)
            if is_live:
                ok, findings_by_area = evaluate_audit(results, lookback)
                audit_areas = AUDIT_AREAS
            else:
                ok, findings_by_area, ready_for_inventory = evaluate_audit_purchasing_only(results, lookback)
                audit_areas = AUDIT_AREAS_PURCHASING
            area_statuses = {}
            for area in AUDIT_AREAS:
                if not is_live and area not in audit_areas:
                    area_statuses[area] = "N/A"
                else:
                    area_statuses[area] = _area_status(findings_by_area.get(area, {"green": [], "yellow": [], "review": [], "red": []}))
            action_plan = get_action_plan(findings_by_area, audit_areas)
            group_folder = _sanitize_folder_name(group_name)
            tenant_results.append({
                "tenant_id": tid, "tenant_name": tname, "area_statuses": area_statuses, "ok": ok, "action_plan": action_plan,
                "results": results, "findings_by_area": findings_by_area, "is_live": is_live, "lookback_days": lookback,
                "scorecard_path": f"{group_folder}/scorecard_{tid}.html",
            })
            html_content = render_scorecard_html(tid, tname, lookback, results, findings_by_area, ok, audit_areas=audit_areas, is_live=is_live, ready_for_inventory_implementation=(ready_for_inventory if not is_live else None))
            (output_dir / f"scorecard_{tid}.html").write_text(html_content, encoding="utf-8")
        group_html = render_group_summary_html(group_name, tenant_results, lookback)
        (output_dir / "index.html").write_text(group_html, encoding="utf-8")
        print(f"Wrote group summary to {output_dir / 'index.html'} and {len(tenant_results)} scorecards to {output_dir}", file=sys.stderr)
        if args.output_aggregate and tenant_results:
            from audit.aggregate import merge_into_aggregate
            merge_into_aggregate(Path(args.output_aggregate), tenant_results, lookback)
            print(f"Merged {len(tenant_results)} tenants into {args.output_aggregate}", file=sys.stderr)
        return 0

    # --- Tenant group mode: run audit for each tenant in the group, then write group summary + per-tenant scorecards ---
    if args.tenant_group and args.from_snowflake and args.html:
        conn = _snowflake_conn()
        try:
            members = get_tenant_group_members(conn, args.tenant_group)
            if not members:
                print(f"No tenants found for group: {args.tenant_group}", file=sys.stderr)
                return 1
            base = Path(args.html).resolve().parent
            output_dir = base / _sanitize_folder_name(args.tenant_group)
            output_dir.mkdir(parents=True, exist_ok=True)
            css_src = Path(__file__).resolve().parent / "css" / "scorecard.css"
            if css_src.is_file():
                (output_dir / "scorecard.css").write_text(css_src.read_text(encoding="utf-8"), encoding="utf-8")
            tenant_results = []
            for i, (tenant_id, tenant_name) in enumerate(members):
                print(f"  [{i+1}/{len(members)}] {tenant_name} ({tenant_id}) ...", file=sys.stderr)
                data = load_results_from_snowflake(tenant_id, args.lookback_days, conn=conn)
                lookback = _int(data.get("lookback_days")) or args.lookback_days
                results = parse_results(data)
                is_live = is_live_with_inventory(results)
                if is_live:
                    ok, findings_by_area = evaluate_audit(results, lookback)
                    audit_areas = AUDIT_AREAS
                else:
                    ok, findings_by_area, ready_for_inventory = evaluate_audit_purchasing_only(results, lookback)
                    audit_areas = AUDIT_AREAS_PURCHASING
                area_statuses = {}
                for area in AUDIT_AREAS:
                    if not is_live and area not in audit_areas:
                        area_statuses[area] = "N/A"
                    else:
                        area_statuses[area] = _area_status(findings_by_area.get(area, {"green": [], "yellow": [], "review": [], "red": []}))
                action_plan = get_action_plan(findings_by_area, audit_areas)
                group_folder = _sanitize_folder_name(args.tenant_group)
                tenant_results.append({
                    "tenant_id": tenant_id,
                    "tenant_name": tenant_name,
                    "area_statuses": area_statuses,
                    "ok": ok,
                    "action_plan": action_plan,
                    "results": results,
                    "findings_by_area": findings_by_area,
                    "is_live": is_live,
                    "lookback_days": lookback,
                    "scorecard_path": f"{group_folder}/scorecard_{tenant_id}.html",
                })
                html_content = render_scorecard_html(tenant_id, tenant_name, lookback, results, findings_by_area, ok, audit_areas=audit_areas, is_live=is_live, ready_for_inventory_implementation=(ready_for_inventory if not is_live else None))
                (output_dir / f"scorecard_{tenant_id}.html").write_text(html_content, encoding="utf-8")
        finally:
            conn.close()
        group_html = render_group_summary_html(args.tenant_group, tenant_results, lookback)
        (output_dir / "index.html").write_text(group_html, encoding="utf-8")
        print(f"Wrote group summary to {output_dir / 'index.html'} and {len(tenant_results)} scorecards to {output_dir}", file=sys.stderr)
        if args.output_aggregate and tenant_results:
            from audit.aggregate import merge_into_aggregate
            merge_into_aggregate(Path(args.output_aggregate), tenant_results, lookback)
            print(f"Merged {len(tenant_results)} tenants into {args.output_aggregate}", file=sys.stderr)
        return 0

    if args.from_snowflake:
        tid = args.tenant_id
        if tid is None:
            print("--from-snowflake requires --tenant-id (or use --tenant-group for a group run).", file=sys.stderr)
            return 1
        data = load_results_from_snowflake(int(tid), args.lookback_days)
        lookback = _int(data.get("lookback_days")) or args.lookback_days
    elif args.from_csv_folder:
        folder = Path(args.from_csv_folder)
        if not folder.is_dir():
            print(f"CSV folder not found: {folder}", file=sys.stderr)
            return 1
        data = load_results_from_csv_folder(folder, args.lookback_days)
        lookback = _int(data.get("lookback_days")) or args.lookback_days
    elif args.from_excel:
        path = Path(args.from_excel)
        if not path.is_file():
            print(f"File not found: {path}", file=sys.stderr)
            return 1
        if path.suffix.lower() == ".csv":
            data = load_results_from_combined_csv(path, args.lookback_days)
        else:
            data = load_results_from_excel(path, args.lookback_days)
        lookback = _int(data.get("lookback_days")) or args.lookback_days
    elif args.from_results:
        path = Path(args.from_results)
        if not path.is_file():
            print(f"Results file not found: {path}", file=sys.stderr)
            return 1
        try:
            raw = json.loads(path.read_text())
        except json.JSONDecodeError as e:
            print(f"Invalid JSON in {path}: {e}", file=sys.stderr)
            return 1
        data = raw.get("results", raw) if isinstance(raw, dict) else raw
        if not isinstance(data, dict):
            print("JSON must be an object (results dict).", file=sys.stderr)
            return 1
        lookback = _int(data.get("lookback_days")) or args.lookback_days

    if data is None:
        parser.print_help()
        print("\nTo run: set tenant_id in sql/00_combined_audit.sql, run in Snowflake, export to Excel/CSV, then use --from-excel. See README.", file=sys.stderr)
        return 1

    # CSV folder: each file has one or more data rows (header already skipped). Use first row per key; pad PO row to 16 cols for parse_results.
    if args.from_csv_folder and data:
        normalized = {"lookback_days": data.get("lookback_days", lookback)}
        for key in RESULT_KEYS:
            rows = data.get(key)
            if not rows or not isinstance(rows, list):
                continue
            row = rows[0] if rows else []
            if isinstance(row, (list, tuple)):
                need = 2 if key == "tenant_info" else (6 if key == "pricebook" else (16 if key == "purchase_orders_summary" else (10 if key == "invoice_materials" else (3 if key == "replenishment_summary" else (9 if key == "returns_summary" else (16 if key == "assessment_data" else (5 if key == "setup_data" else (8 if key == "usage_checks" else (1 if key == "inventory_settings" else 4)))))))))
                normalized[key] = [list(row) + [None] * max(0, need - len(row))]
            else:
                normalized[key] = [row]
        data = normalized

    results = parse_results(data)
    is_live = is_live_with_inventory(results)
    ready_for_inventory = None
    if is_live:
        ok, findings_by_area = evaluate_audit(results, lookback)
        audit_areas = AUDIT_AREAS
    else:
        ok, findings_by_area, ready_for_inventory = evaluate_audit_purchasing_only(results, lookback)
        audit_areas = AUDIT_AREAS_PURCHASING

    # --- Output: tenant, scorecard, Settings, Setup, then findings by area ---
    tenant_id = args.tenant_id if args.tenant_id is not None else results.get("tenant_id")
    tenant_name = args.tenant_name if args.tenant_name else results.get("tenant_name")

    if args.html:
        base = Path(args.html).resolve().parent
        folder_name = f"scorecard_{tenant_id or 0}_{_sanitize_folder_name(tenant_name or 'tenant')}"
        output_dir = base / folder_name
        output_dir.mkdir(parents=True, exist_ok=True)
        html_content = render_scorecard_html(tenant_id, tenant_name, lookback, results, findings_by_area, ok, audit_areas=audit_areas, is_live=is_live, ready_for_inventory_implementation=(ready_for_inventory if not is_live else None))
        (output_dir / "scorecard.html").write_text(html_content, encoding="utf-8")
        css_src = Path(__file__).resolve().parent / "css" / "scorecard.css"
        if css_src.is_file():
            (output_dir / "scorecard.css").write_text(css_src.read_text(encoding="utf-8"), encoding="utf-8")
            print(f"Wrote scorecard to {output_dir / 'scorecard.html'} and CSS to {output_dir}", file=sys.stderr)
        else:
            print(f"Wrote scorecard to {output_dir / 'scorecard.html'} (no css/scorecard.css found)", file=sys.stderr)

    if args.output_aggregate and data is not None:
        action_plan = get_action_plan(findings_by_area, audit_areas)
        area_statuses_single = {}
        for area in AUDIT_AREAS:
            if not is_live and area not in audit_areas:
                area_statuses_single[area] = "N/A"
            else:
                area_statuses_single[area] = _area_status(findings_by_area.get(area, {"green": [], "yellow": [], "review": [], "red": []}))
        scorecard_path = None
        if args.html:
            folder_name = f"scorecard_{tenant_id or 0}_{_sanitize_folder_name(tenant_name or 'tenant')}"
            scorecard_path = f"{folder_name}/scorecard.html"
        from audit.aggregate import merge_into_aggregate
        merge_into_aggregate(Path(args.output_aggregate), [{
            "tenant_id": tenant_id, "tenant_name": tenant_name, "lookback_days": lookback,
            "is_live": is_live, "results": results, "findings_by_area": findings_by_area,
            "area_statuses": area_statuses_single, "action_plan": action_plan,
            "scorecard_path": scorecard_path,
        }], lookback)
        print(f"Merged 1 tenant into {args.output_aggregate}", file=sys.stderr)

    width = 62
    print("=" * width)
    print("  INVENTORY SETUP / USAGE AUDIT — CUSTOMER SCORECARD")
    print("=" * width)
    if tenant_id is not None:
        print(f"  Tenant ID:   {tenant_id}")
    if tenant_name:
        print(f"  Tenant:      {tenant_name}")
    print(f"  Lookback:    {lookback} days")
    print()

    # --- Scorecard: one row per area with status ---
    if not is_live:
        print("  (Purchasing audit — not yet live with inventory)")
    print("  SCORECARD BY AREA")
    print("  " + "-" * (width - 2))
    status_symbol = {"Red": "[!]", "Review": "[~]", "Yellow": "[~]", "Green": "[OK]"}
    for area in audit_areas:
        findings = findings_by_area.get(area, {"green": [], "yellow": [], "red": []})
        status = _area_status(findings)
        sym = status_symbol.get(status, "•")
        print(f"  {sym}  {area:<28}  {status}")
    print("  " + "-" * (width - 2))
    overall = "No critical issues" if ok else "Address red items to improve setup and usage"
    print(f"  Overall: {overall}")
    if not is_live:
        if _is_preparing_for_go_live(results, is_live):
            go_live_ready, go_live_needs = _go_live_readiness(results)
            if go_live_ready:
                print("  Go Live Readiness: Ready to set your go live date and do inventory beginning balance counts.")
            else:
                print("  Go Live Readiness:")
                for n in go_live_needs:
                    print(f"    • {n}")
        else:
            if ready_for_inventory:
                print("  Inventory Readiness: Ready to start inventory implementation.")
            else:
                areas_needing_work = _readiness_areas_needing_work(findings_by_area)
                areas_str = _format_areas_for_message(areas_needing_work)
                if areas_str:
                    print(f"  Inventory Readiness: Address red items in {areas_str} to prepare for inventory implementation.")
                else:
                    print("  Inventory Readiness: Address red items above to prepare for inventory implementation.")
    print()

    print("  --- Settings ---")
    cfg = results.get("readiness_config") or {}
    fg = results.get("feature_gates") or {}
    print("    Feature gates:")
    print(f"      Purchasing module: {'On' if cfg.get('purchasing_module_on') else 'Off' if cfg.get('purchasing_module_on') is False else '—'}")
    print(f"      Inventory module:  {'On' if cfg.get('inventory_module_on') else 'Off' if cfg.get('inventory_module_on') is False else '—'}")
    if fg:
        print(f"      Item requisitions: {'On' if fg.get('item_requisitions') else 'Off'}")
        print(f"      Requisition - Closeout: {'On' if fg.get('requisition_closeout') else 'Off'}")
        print(f"      Enable Requisition Workflow for Service Job: {'On' if fg.get('install_requisitions') else 'Off'}")
        print(f"      Transfers to jobs: {'On' if fg.get('transfers_to_jobs') else 'Off'}")
        print(f"      Consignment: {'On' if fg.get('consignment') else 'Off'}")
        print(f"      Purchase approval workflow: {'On' if fg.get('purchase_approval_workflow') else 'Off'}")
        print(f"      Granular WAC: {'On' if fg.get('granular_wac') else 'Off'}")
        if fg.get('stock_optimizer') is not None:
            print(f"      Stock Optimizer: {'On' if fg.get('stock_optimizer') else 'Off'}")
    print()
    if is_live:
        print("    Inventory Settings (in UI) — customer-controllable:")
        inv_settings = results.get("inventory_settings")
        if inv_settings:
            for line in format_inventory_settings(inv_settings):
                print(line)
        else:
            print("      (not found or not available)")
        print()
    else:
        inv_settings = results.get("inventory_settings")
        if inv_settings:
            start = inv_settings.get("BeginningDate") or inv_settings.get("beginningDate")
            if start is not None:
                print(f"    Inventory tracking start date: {str(start).strip()[:10]}")
        print()
    if is_live:
        print("  --- Setup (counts) ---")
        isinv = results.get("isinventory_counts") or {}
        active_mats = isinv.get("active_materials_count")
        mat_isinv = isinv.get("materials_isinventory_count")
        uom_count = results.get("materials_with_unit_of_measure", 0)
        active_eq = isinv.get("active_equipment_count")
        eq_isinv = isinv.get("equipment_isinventory_count")
        eq_serialized = results.get("equipment_serialized_count", 0)
        print(f"    Materials: total {active_mats if active_mats is not None else '—'}; with IsInventory: {mat_isinv if mat_isinv is not None else '—'}; with unit of measure: {uom_count}")
        print(f"    Equipment: total {active_eq if active_eq is not None else '—'}; with IsInventory: {eq_isinv if eq_isinv is not None else '—'}; serialized: {eq_serialized}")
        setup = results.get("setup_data")
        if setup and isinstance(setup, dict):
            truck_total = setup.get("truck_total", 0)
            truck_with_tpl = setup.get("truck_with_template", 0)
            wh_total = setup.get("warehouse_total", 0)
            wh_with_tpl = setup.get("warehouse_with_template", 0)
            templates_under = setup.get("templates_under_20_active_items", 0)
            truck_pct = (100 * truck_with_tpl // truck_total) if truck_total > 0 else 0
            truck_ok = truck_total == 0 or truck_pct >= TRUCK_TEMPLATE_PCT_MIN
            wh_ok = wh_total == 0 or wh_with_tpl >= WAREHOUSE_WITH_TEMPLATE_MIN
            templates_ok = templates_under == 0
            print(f"    {'[OK]' if truck_ok else '[!!]'} Trucks: {truck_total} total (active), {truck_with_tpl} with inventory template (min 80% expected)")
            print(f"    {'[OK]' if wh_ok else '[!!]'} Warehouses: {wh_total} total (active), {wh_with_tpl} with inventory template (at least 1 expected)")
            print(f"    {'[OK]' if templates_ok else '[!!]'} Templates with < {TEMPLATE_MIN_ACTIVE_ITEMS} active items: {templates_under}")
        print()

    # --- Findings by area ---
    print("  --- Findings by area ---")
    for area in audit_areas:
        findings = findings_by_area.get(area, {"green": [], "yellow": [], "review": [], "red": []})
        if not any(findings.values()):
            continue
        status = _area_status(findings)
        sym = status_symbol.get(status, "•")
        print(f"    [{sym} {area}]")
        for g in findings["green"]:
            print(f"      [OK]    {g}")
        for y in findings["yellow"]:
            print(f"      [WARN]  {y}")
        for rev in findings.get("review") or []:
            print(f"      [REV]   {rev}")
        for r in findings["red"]:
            print(f"      [!!]    {r}")
        print()

    print("  " + "=" * (width - 2))
    if ok:
        print("  Audit: No critical issues. Review warnings as needed.")
    else:
        print("  Audit: Address red items to improve setup and usage.")
    print("=" * width)
    return 0 if ok else 2


if __name__ == "__main__":
    sys.exit(main())
