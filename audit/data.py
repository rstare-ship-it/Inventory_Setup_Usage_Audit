"""
Data loading: all functions that read raw audit data from CSV, Excel, Snowflake, or row lists
and return a normalised data dict for parse_results().
"""
from __future__ import annotations

import csv
import os
import sys
from pathlib import Path

from audit.constants import RESULT_KEYS


def _int(v) -> int:
    """Coerce to int (handles JSON strings, Snowflake results, None)."""
    if v is None:
        return 0
    if isinstance(v, int):
        return v
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def data_from_audit_rows(rows: list, lookback_days: int = 90) -> dict:
    """Build a data dict from audit query rows.
    Each row must be [source, v1, v2, ..., v16] (e.g. from MCP or CSV)."""
    data: dict = {"lookback_days": lookback_days}
    for row in rows:
        if not row or len(row) < 1:
            continue
        source = (row[0] or "").strip().lower()
        key = next((k for k in RESULT_KEYS if k.lower() == source), None)
        if key:
            values = list(row[1:17])
            if len(values) < 16:
                values.extend([None] * (16 - len(values)))
            data[key] = [values[:16]]
    return data


def load_results_from_combined_csv(path: Path, lookback_days: int = 90) -> dict:
    """Read combined audit result from a single CSV (header: source, v1..v16; 9 data rows)."""
    data: dict = {"lookback_days": lookback_days}
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if len(rows) < 2:
        return data
    header = [str(x).strip().lower() if x else "" for x in rows[0]]
    if not header or header[0] != "source":
        return data
    col_index = {f"v{i}": i for i in range(1, 17)}
    for j, h in enumerate(header[1:], start=1):
        if h and h in col_index:
            col_index[h] = j
    known = {s.lower() for s in RESULT_KEYS}
    for r in rows[1:]:
        if not r:
            continue
        src = (r[0] or "").strip()
        if isinstance(src, (int, float)):
            src = str(int(src))
        src_lower = src.lower() if src else ""
        if src_lower in known:
            key = next(k for k in RESULT_KEYS if k.lower() == src_lower)
            values = [None] * 16
            for vi in range(1, 17):
                vkey = f"v{vi}"
                j = col_index.get(vkey, vi)
                if j < len(r):
                    values[vi - 1] = r[j]
            data[key] = [values]
    return data


def _excel_cell_value(cell) -> object:
    v = cell.value
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10] if hasattr(v, "date") else str(v)
    return v


def load_results_from_excel(path: Path, lookback_days: int = 90) -> dict:
    """Read combined audit result from one Excel sheet (source + v1..v16, 9 rows)."""
    try:
        import openpyxl
    except ImportError:
        print("--from-excel requires openpyxl.  pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    data: dict = {"lookback_days": lookback_days}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if len(wb.sheetnames) >= 1:
            ws = wb.active
            rows = [[_excel_cell_value(c) for c in row] for row in ws.iter_rows()]
            if len(rows) >= 2 and rows[0]:
                header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
                if header and header[0] == "source":
                    known = {s.lower() for s in RESULT_KEYS}
                    for r in rows[1:]:
                        if not r:
                            continue
                        src = (r[0] or "").strip()
                        if isinstance(src, (int, float)):
                            src = str(int(src))
                        if src.lower() in known:
                            key = next(k for k in RESULT_KEYS if k.lower() == src.lower())
                            values = list(r[1:17])
                            if len(values) < 16:
                                values = values + [None] * (16 - len(values))
                            data[key] = [values[:16]]
    finally:
        wb.close()
    return data


def load_results_from_csv_folder(folder: Path, lookback_days: int = 90) -> dict:
    """Read from per-key CSVs: e.g. purchase_orders_summary.csv, invoice_materials.csv, …"""
    data: dict = {"lookback_days": lookback_days}
    for key in RESULT_KEYS:
        path = folder / f"{key}.csv"
        if not path.is_file():
            continue
        rows = []
        try:
            with path.open(newline="", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                first = True
                for row in reader:
                    if first:
                        first = False
                        continue
                    rows.append(row)
        except Exception as e:
            print(f"Warning: could not read {path}: {e}", file=sys.stderr)
            continue
        if rows:
            data[key] = rows
    return data


def _snowflake_conn():
    """Create a Snowflake connection (caller must close).
    Requires SNOWFLAKE_ACCOUNT and SNOWFLAKE_USER env vars."""
    try:
        import snowflake.connector
    except ImportError:
        print(
            "--from-snowflake requires snowflake-connector-python.  "
            "pip install snowflake-connector-python",
            file=sys.stderr,
        )
        sys.exit(1)
    account = os.environ.get("SNOWFLAKE_ACCOUNT")
    user = os.environ.get("SNOWFLAKE_USER")
    if not account or not user:
        print(
            "Set SNOWFLAKE_ACCOUNT and SNOWFLAKE_USER in the environment for --from-snowflake.",
            file=sys.stderr,
        )
        sys.exit(1)
    region = os.environ.get("SNOWFLAKE_REGION")
    if region and not account.endswith(f".{region}"):
        account = f"{account}.{region}"
    conn_params: dict = {
        "account": account,
        "user": user,
        "authenticator": os.environ.get("SNOWFLAKE_AUTHENTICATOR", "externalbrowser"),
    }
    for key, env_key in (
        ("warehouse", "SNOWFLAKE_WAREHOUSE"),
        ("database", "SNOWFLAKE_DATABASE"),
        ("schema", "SNOWFLAKE_SCHEMA"),
        ("role", "SNOWFLAKE_ROLE"),
    ):
        val = os.environ.get(env_key)
        if val:
            conn_params[key] = val
    return snowflake.connector.connect(**conn_params)


def get_tenant_group_members(conn, group_name: str) -> list[tuple[int, str]]:
    """Return list of (tenant_id, tenant_name) for all tenants in the given group. Uses MASTER_DB."""
    query = """
    SELECT r._TENANT_ID, TRIM(COALESCE(r._TENANT_NAME, ''))
    FROM tenant_data.MASTER_DB.TENANTGROUP g
    JOIN tenant_data.MASTER_DB.TENANTRECORD_GROUPS_TENANTGROUP j ON g.ID = j.TENANTGROUP
    JOIN tenant_data.MASTER_DB.TENANTRECORD r ON j.TENANTRECORD = r._TENANT_ID
    WHERE g.NAME = %s
    ORDER BY r._TENANT_NAME
    """
    with conn.cursor() as cur:
        cur.execute(query, (group_name,))
        rows = cur.fetchall()
    return [
        (_int(row[0]), (row[1] or "").strip() or f"Tenant {row[0]}")
        for row in rows
        if row and row[0] is not None
    ]


def _read_sql_file(sql_path: Path | None = None) -> str:
    """Read and validate the combined audit SQL template."""
    sql_file = sql_path or (Path(__file__).resolve().parent.parent / "sql" / "00_combined_audit.sql")
    if not sql_file.is_file():
        print(f"SQL file not found: {sql_file}", file=sys.stderr)
        sys.exit(1)
    raw_sql = sql_file.read_text(encoding="utf-8")
    if "SELECT 0 AS tenant_id" not in raw_sql:
        print(
            "SQL file missing expected tenant_param placeholder (SELECT 0 AS tenant_id).",
            file=sys.stderr,
        )
        sys.exit(1)
    return raw_sql


def _build_tenant_sql(raw_sql: str, tenant_id: int) -> str:
    """Inject a specific tenant_id into the SQL template."""
    return raw_sql.replace("SELECT 0 AS tenant_id", f"SELECT {int(tenant_id)} AS tenant_id")


def load_results_from_snowflake(
    tenant_id: int,
    lookback_days: int = 90,
    sql_path: Path | None = None,
    conn=None,
) -> dict:
    """Run the combined audit query in Snowflake and return a data dict for parse_results.
    Pass `conn` to reuse an existing connection (it will not be closed)."""
    raw_sql = _read_sql_file(sql_path)
    sql = _build_tenant_sql(raw_sql, tenant_id)

    own_conn = conn is None
    if own_conn:
        conn = _snowflake_conn()
    data: dict = {"lookback_days": lookback_days}
    try:
        with conn.cursor() as cur:
            cur.execute(sql)
            rows = cur.fetchall()
        for row in rows:
            if not row or len(row) < 1:
                continue
            source = (row[0] or "").strip().lower()
            key = next((k for k in RESULT_KEYS if k.lower() == source), None)
            if key:
                values = list(row[1:17])
                if len(values) < 16:
                    values.extend([None] * (16 - len(values)))
                data[key] = [values[:16]]
    finally:
        if own_conn:
            conn.close()
    return data


def load_results_batch_from_snowflake(
    tenant_ids: list[int],
    lookback_days: int = 90,
    sql_path: Path | None = None,
    conn=None,
) -> dict[int, dict]:
    """Run the combined audit query for multiple tenants, reusing one connection.
    Returns {tenant_id: data_dict}. Pass `conn` to reuse an existing connection.

    Runs one query per tenant sequentially on the shared connection so only one
    Okta authentication is needed for the whole group.
    """
    if not tenant_ids:
        return {}

    raw_sql = _read_sql_file(sql_path)

    own_conn = conn is None
    if own_conn:
        conn = _snowflake_conn()

    results: dict[int, dict] = {}
    try:
        for tid in tenant_ids:
            sql = _build_tenant_sql(raw_sql, tid)
            data: dict = {"lookback_days": lookback_days}
            with conn.cursor() as cur:
                cur.execute(sql)
                rows = cur.fetchall()
            for row in rows:
                if not row or len(row) < 1:
                    continue
                source = (row[0] or "").strip().lower()
                key = next((k for k in RESULT_KEYS if k.lower() == source), None)
                if key:
                    values = list(row[1:17])
                    if len(values) < 16:
                        values.extend([None] * (16 - len(values)))
                    data[key] = [values[:16]]
            results[tid] = data
    finally:
        if own_conn:
            conn.close()
    return results
